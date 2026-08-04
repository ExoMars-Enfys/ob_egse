#!/usr/bin/env python3
"""Recursively summarise a thermal-test folder into a new Excel workbook.

This script deliberately reuses ``utility_modules.eb_packet_utility`` from the
Enfys analysis project, so packet decoding and engineering-unit conversions stay
consistent with analysis.py / plot_all.py.  A single selected folder supplies
RS422, PSU, MSC1, MSC2, and TP1000/TRP logs; the workbook is created beside them.
"""

from __future__ import annotations

import csv
import math
import posixpath
import re
import statistics
import sys
import tkinter as tk
import warnings
import zipfile
from dataclasses import dataclass, field
from datetime import date, datetime, time, timedelta
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import Any, Iterable
from xml.etree import ElementTree

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo

RS422_NAME_RE = re.compile(r"^RS422if_.*\.log$", re.IGNORECASE)
EXTRACTOR_VERSION = "2026.07.31-single-folder-workflow"
FIXED_STATE_CURRENT_WINDOW_S = 5.0
STATE6_CURRENT_TRIGGER_S = 150.0
NAME_TS_RE = re.compile(r"(\d{4}-\d{2}-\d{2})_(\d{2})-(\d{2})-(\d{2})")
FULL_TS_RE = re.compile(r"(\d{4}[-/]\d{2}[-/]\d{2})[ _T](\d{2})[:-](\d{2})[:-](\d{2})(?:[.,](\d{1,6}))?")
TIME_ONLY_RE = re.compile(r"\b(\d{2}):(\d{2}):(\d{2})(?:[.,](\d{1,6}))?\b")
HEX_WORD_RE = re.compile(r"\b[0-9A-Fa-f]{4}\b")

BASE_HEADERS = [
    "Temp (°C)",
    "run",
    "SFT Date/Time",
    "EB Chamber Temp (°C)",
    "OB Chamber Temp(°C)",
    "ROV_THERM (°C)",
    "CH4_I at OB Switch on (A)",
    "DIG Temp (°C)",
    "DETEC Temp (°C)",
    "MECH Temp (°C)",
    "MTR Temp (°C)",
    "OB +3V3 (V)",
    "OB +3V3 Max(V)",
    "+3V3 Max Delta",
    "OB +1V5 (V)",
    "OB +1V5 Max (V)",
    "+1V5 Max Delta",
    "Sci Packets",
    "HTSNK Start Temp(adu)",
    "HTSNK End Temp(adu)",
    "SWIR Start Temp (adu)",
    "SWIR End Temp (adu)",
    "MWIR Start Temp (adu)",
    "MWIR End Temp (adu)",
    "ROV Dwell Before CH3/CH4 (min)",
    "ROV Dwell >10 min",
    "Matrix EB Setpoint (°C)",
    "Matrix OB Setpoint (°C)",
    "Backfilled From",
]

STATE_CHECK_LABELS = (
    "State1",
    "Standby",
    "State2",
    "State3",
    "State4",
    "State5",
    "State6 Initial",
    "State6 Heaters ON",
    "State6 Heaters OFF",
    "State7",
)
STATE_CHECK_HEADERS = [f"{label} Current (mA)" for label in STATE_CHECK_LABELS]
HEADERS = BASE_HEADERS + STATE_CHECK_HEADERS

# Preference order is selected/latter SFT first, then valid earlier SFTs that
# may fill blank measurements. SFT9 is +10/+10 only. SFT10 is excluded.
TEST_MATRIX_GROUPS: tuple[tuple[float, float, tuple[int, ...]], ...] = (
    (-40, -50, (25, 24)),
    (-40, -40, (8,)),
    (-40, -20, (13,)),
    (-40, 0, (23,)),
    (-30, -30, (7,)),
    (-20, -50, (26,)),
    (-20, -40, (18,)),
    (-20, -20, (17, 6)),
    (-20, 0, (16,)),
    (-20, 20, (22,)),
    (-10, -10, (5, 4, 3)),
    (0, -50, (27,)),
    (0, -40, (19,)),
    (0, -20, (11,)),
    (0, 0, (15, 2)),
    (0, 20, (21,)),
    (10, 10, (9,)),
    (20, -40, (20,)),
    (20, -20, (29,)),
    (20, 0, (28,)),
    (20, 20, (14,)),
)

META_COLUMNS = {
    "temp": "Temp (°C)",
    "run": "run",
    "eb_chamber_temp": "EB Chamber Temp (°C)",
    "ob_chamber_temp": "OB Chamber Temp(°C)",
    "rov_therm": "ROV_THERM (°C)",
}

SCIENCE_TEMP_FIELDS = {
    "HTSNK_START": ("HEATSINK_START_TEMP", "HTSNK_START_TEMP", "HTSINK_START_TEMP"),
    "HTSNK_END": ("HEATSINK_END_TEMP", "HTSNK_END_TEMP", "HTSINK_END_TEMP"),
    "SWIR_START": ("SWIR_START_TEMP",),
    "SWIR_END": ("SWIR_END_TEMP",),
    "MWIR_START": ("MWIR_START_TEMP",),
    "MWIR_END": ("MWIR_END_TEMP",),
}

SCIENCE_OUTPUT_HEADERS = {
    "HTSNK_START": "HTSNK Start Temp(adu)",
    "HTSNK_END": "HTSNK End Temp(adu)",
    "SWIR_START": "SWIR Start Temp (adu)",
    "SWIR_END": "SWIR End Temp (adu)",
    "MWIR_START": "MWIR Start Temp (adu)",
    "MWIR_END": "MWIR End Temp (adu)",
}

OB_FIRST_TEMPERATURE_HEADERS = {
    "DIG Temp (°C)",
    "DETEC Temp (°C)",
    "MECH Temp (°C)",
    "MTR Temp (°C)",
}


@dataclass
class Packet:
    time: datetime | None
    hk: Any


@dataclass
class Summary:
    values: dict[str, Any]
    hk_count: int
    science_packets: int
    first_time: datetime | None
    last_time: datetime | None
    ob_switch_time: datetime | None
    warnings: list[str]
    state_condition_times: dict[str, datetime] = field(default_factory=dict)


@dataclass(frozen=True)
class TemperaturePoint:
    time: datetime
    value: float
    source: Path


def finite(value: Any) -> float | None:
    try:
        number = float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def choose(values: Iterable[Any], mode: str) -> float | None:
    nums = [n for value in values if (n := finite(value)) is not None]
    if not nums:
        return None
    if mode == "first":
        return nums[0]
    if mode == "last":
        return nums[-1]
    if mode == "mean":
        return statistics.fmean(nums)
    return statistics.median(nums)


def _hk_operating_state(hk: Any) -> int | None:
    value = finite(getattr(hk, "CURRENT_OPERATING_STATE", None))
    return int(value) if value is not None else None


def _hk_flag(hk: Any, group_name: str, flag_name: str) -> bool:
    group = getattr(hk, group_name, None)
    return bool(getattr(group, flag_name, 0)) if group is not None else False


def _hk_tec_current_a(hk: Any) -> float | None:
    raw = finite(getattr(hk, "EB_TEC_DRIVE_CURRENT", None))
    return None if raw is None else raw * 0.0000162


def _hk_heaters_commanded(hk: Any) -> bool:
    thrm = getattr(hk, "THRM_STATUS", None)
    if thrm is None:
        return False
    mech_commanded = bool(getattr(thrm, "MM", 0) or getattr(thrm, "MA", 0))
    det_commanded = bool(getattr(thrm, "DM", 0) or getattr(thrm, "DA", 0))
    return mech_commanded and det_commanded


def _hk_heater_outputs_off(hk: Any) -> bool:
    thrm = getattr(hk, "THRM_STATUS", None)
    return bool(thrm is not None and not getattr(thrm, "HMS", 0) and not getattr(thrm, "HDS", 0))


def _hk_boards_enabled(hk: Any) -> bool:
    return _hk_flag(hk, "INSTR_STATUS_FLAGS", "OB_MECHANISM_BOARD_ENABLED") and _hk_flag(
        hk, "INSTR_STATUS_FLAGS", "OB_DETECTOR_BOARD_ENABLED"
    )


def _hk_motor_moving(hk: Any) -> bool:
    return _hk_flag(hk, "MTR_FLAGS", "MOVING")


def _hk_heater_config_matches(hk: Any, maximum: int, minimum: int) -> bool | None:
    field_pairs = (
        ("OB_THERMAL_MECH_MAX", maximum),
        ("OB_THERMAL_MECH_MIN", minimum),
        ("OB_THERMAL_DET_MAX", maximum),
        ("OB_THERMAL_DET_MIN", minimum),
    )
    observed: list[bool] = []
    for name, expected in field_pairs:
        value = finite(getattr(hk, name, None))
        if value is not None:
            observed.append(int(value) == expected)
    return all(observed) if observed else None


def _acquisition_intervals(
    points: list[Packet],
) -> list[tuple[datetime, datetime | None, datetime | None]]:
    """Return acquisition start/end and the t+150 s state-current check point."""
    intervals: list[tuple[datetime, datetime | None, datetime | None]] = []
    start_index: int | None = None
    for index, point in enumerate(points):
        in_acquisition = _hk_operating_state(point.hk) == 0x08
        if in_acquisition and start_index is None:
            start_index = index
        if not in_acquisition and start_index is not None:
            start_point = points[start_index]
            if start_point.time is not None:
                target = start_point.time + timedelta(seconds=STATE6_CURRENT_TRIGGER_S)
                check_time = next(
                    (
                        candidate.time
                        for candidate in points[start_index:index]
                        if candidate.time is not None and candidate.time >= target
                    ),
                    None,
                )
                intervals.append((start_point.time, point.time, check_time))
            start_index = None
    if start_index is not None:
        start_point = points[start_index]
        if start_point.time is not None:
            target = start_point.time + timedelta(seconds=STATE6_CURRENT_TRIGGER_S)
            check_time = next(
                (
                    candidate.time
                    for candidate in points[start_index:]
                    if candidate.time is not None and candidate.time >= target
                ),
                None,
            )
            intervals.append((start_point.time, None, check_time))
    return intervals


def detect_state_condition_times(hk_packets: list[Packet]) -> dict[str, datetime]:
    """Reconstruct FFT state-check points directly from decoded HK conditions."""
    points = sorted(
        (packet for packet in hk_packets if packet.time is not None),
        key=lambda packet: packet.time or datetime.min,
    )
    if not points:
        return {}

    def first_matching(
        predicate: Any,
        *,
        after: datetime | None = None,
        before: datetime | None = None,
    ) -> datetime | None:
        for point in points:
            if point.time is None:
                continue
            if after is not None and point.time <= after:
                continue
            if before is not None and point.time >= before:
                break
            if predicate(point.hk):
                return point.time
        return None

    intervals = _acquisition_intervals(points)
    first_acq_start = intervals[0][0] if intervals else None
    second_acq_start = intervals[1][0] if len(intervals) > 1 else None

    result: dict[str, datetime] = {}
    state1 = first_matching(lambda hk: _hk_operating_state(hk) == 0x02)
    if state1 is not None:
        result["State1"] = state1
    standby = first_matching(
        lambda hk: _hk_operating_state(hk) == 0x04,
        after=state1,
    )
    if standby is not None:
        result["Standby"] = standby

    state2 = first_matching(
        lambda hk: (
            _hk_operating_state(hk) == 0x04
            and _hk_heaters_commanded(hk)
            and not _hk_boards_enabled(hk)
            and (_hk_tec_current_a(hk) or 0.0) < 0.2
        ),
        after=standby,
        before=first_acq_start,
    )
    if state2 is not None:
        result["State2"] = state2

    acquisition_labels = (
        "State6 Initial",
        "State6 Heaters ON",
        "State6 Heaters OFF",
    )
    for label, (_start, _end, check_time) in zip(acquisition_labels, intervals):
        if check_time is not None:
            result[label] = check_time

    post_initial_acq = intervals[0][1] or intervals[0][2] or intervals[0][0] if intervals else state2

    def state3_configuration(hk: Any) -> bool:
        match = _hk_heater_config_matches(hk, 0x08D3, 0x08A3)
        return True if match is None else match

    def flight_configuration(hk: Any) -> bool:
        match = _hk_heater_config_matches(hk, 0x079A, 0x0738)
        return _hk_heater_outputs_off(hk) if match is None else match

    state3 = first_matching(
        lambda hk: (
            _hk_operating_state(hk) == 0x04
            and _hk_boards_enabled(hk)
            and (_hk_tec_current_a(hk) or 0.0) < 0.2
            and state3_configuration(hk)
        ),
        after=post_initial_acq,
        before=second_acq_start,
    )
    if state3 is not None:
        result["State3"] = state3

    state4 = first_matching(
        lambda hk: (
            _hk_operating_state(hk) == 0x04
            and _hk_boards_enabled(hk)
            and (_hk_tec_current_a(hk) or 0.0) > 1.0
            and state3_configuration(hk)
        ),
        after=state3,
        before=second_acq_start,
    )
    if state4 is not None:
        result["State4"] = state4

    state7 = first_matching(
        lambda hk: (
            _hk_operating_state(hk) == 0x04
            and _hk_boards_enabled(hk)
            and (_hk_tec_current_a(hk) or 0.0) > 1.0
            and _hk_motor_moving(hk)
            and state3_configuration(hk)
        ),
        after=state4,
        before=second_acq_start,
    )
    if state7 is not None:
        result["State7"] = state7

    state5 = first_matching(
        lambda hk: (
            _hk_operating_state(hk) == 0x04
            and _hk_boards_enabled(hk)
            and (_hk_tec_current_a(hk) or 0.0) > 1.0
            and not _hk_motor_moving(hk)
            and flight_configuration(hk)
        ),
        after=state7 or state4,
        before=second_acq_start,
    )
    if state5 is not None:
        result["State5"] = state5
    return result


def parse_filename_time(path: Path) -> datetime | None:
    match = NAME_TS_RE.search(path.name)
    if not match:
        return None
    return datetime.strptime(
        f"{match.group(1)} {match.group(2)}:{match.group(3)}:{match.group(4)}",
        "%Y-%m-%d %H:%M:%S",
    )


def parse_bound(text: str | None, *, end: bool = False) -> datetime | None:
    if not text:
        return None
    for fmt in ("%Y-%m-%d_%H-%M-%S", "%Y-%m-%dT%H:%M:%S", "%Y-%m-%d %H:%M:%S", "%Y-%m-%d"):
        try:
            result = datetime.strptime(text, fmt)
            if end and fmt == "%Y-%m-%d":
                return result.replace(hour=23, minute=59, second=59, microsecond=999999)
            return result
        except ValueError:
            pass
    raise ValueError(f"Unsupported date/time: {text!r}")


def discover_logs(root: Path, start: datetime | None, end: datetime | None) -> list[Path]:
    logs: list[Path] = []
    for path in root.rglob("*"):
        if not path.is_file() or not RS422_NAME_RE.match(path.name):
            continue
        stamp = parse_filename_time(path) or datetime.fromtimestamp(path.stat().st_mtime)
        if start and stamp < start:
            continue
        if end and stamp > end:
            continue
        logs.append(path)
    return sorted(logs, key=lambda p: (parse_filename_time(p) or datetime.min, str(p).lower()))


def parse_context_timestamp(lines: list[str], index: int, known_date: datetime | None) -> datetime | None:
    for pos in range(index, max(-1, index - 5), -1):
        text = lines[pos].strip()
        direct = NAME_TS_RE.fullmatch(text)
        if direct:
            return parse_filename_time(Path(text + ".log"))
        match = FULL_TS_RE.search(text)
        if match:
            micros = (match.group(5) or "").ljust(6, "0")
            normalized = (
                f"{match.group(1).replace('/', '-')} {match.group(2)}:{match.group(3)}:{match.group(4)}.{micros}"
            )
            return datetime.strptime(normalized, "%Y-%m-%d %H:%M:%S.%f")
        match = TIME_ONLY_RE.search(text)
        if match and known_date:
            micros = (match.group(4) or "").ljust(6, "0")
            return known_date.replace(
                hour=int(match.group(1)),
                minute=int(match.group(2)),
                second=int(match.group(3)),
                microsecond=int(micros or 0),
            )
    return None


def payloads_from_log(path: Path) -> list[tuple[datetime | None, bytes, int]]:
    lines = path.read_text(encoding="utf-8", errors="ignore").splitlines()
    output: list[tuple[datetime | None, bytes, int]] = []
    known_time = parse_filename_time(path)

    telemetry_indices = [i for i, line in enumerate(lines) if "Telemetry Data:" in line]
    if telemetry_indices:
        for index in telemetry_indices:
            if index + 1 >= len(lines):
                continue
            try:
                payload = bytes(int(token, 16) for token in lines[index + 1].strip().split())
            except ValueError:
                continue
            if len(payload) < 6:
                continue
            timestamp = parse_context_timestamp(lines, index, known_time)
            if timestamp:
                known_time = timestamp
            output.append((timestamp, payload, (payload[5] >> 2) & 0x3F))
        return output

    for line in lines:
        parts = line.split(" - ", 1)
        if len(parts) != 2:
            continue
        words = HEX_WORD_RE.findall(parts[1])
        if not words:
            continue
        try:
            payload = bytes.fromhex("".join(words))
        except ValueError:
            continue
        if len(payload) < 6:
            continue
        timestamp = parse_context_timestamp([parts[0]], 0, known_time)
        if timestamp:
            known_time = timestamp
        output.append((timestamp, payload, (payload[5] >> 2) & 0x3F))
    return output


def import_decoder(project_root: Path | None) -> Any:
    candidates = []
    if project_root:
        candidates.extend([project_root, project_root / "src"])
    candidates.extend([Path(__file__).resolve().parent, Path(__file__).resolve().parent.parent])
    for candidate in candidates:
        if (candidate / "utility_modules").is_dir():
            sys.path.insert(0, str(candidate))
            break
    try:
        from utility_modules import eb_packet_utility  # type: ignore
    except ImportError as exc:
        raise SystemExit(
            "Could not import utility_modules.eb_packet_utility. "
            "Run this script inside the Enfys analysis project or pass --project-root."
        ) from exc
    return eb_packet_utility


def convert_hk(hk: Any, decoder: Any) -> dict[str, float | None]:
    thermistor = getattr(decoder, "thermistor_adu_to_temp", None) or getattr(decoder, "decode_eb_trps", None)
    ob_thermistor = getattr(decoder, "decode_ob_trps", None) or getattr(decoder, "adu_to_temp", None)

    def attr(name: str) -> float | None:
        return finite(getattr(hk, name, None))

    def converted(function: Any, name: str) -> float | None:
        raw = getattr(hk, name, None)
        if raw is None or not callable(function):
            return None
        try:
            return finite(function(raw))
        except Exception:
            return None

    mcu = attr("EB_MCU_INTERNAL_TEMP")
    peltier = attr("EB_PELTIER_TEMP")
    result = {
        "MCU temp (°C)": None if mcu is None else mcu * 0.01637198 - 273.0,
        "Peltier Temp(°C)": None if peltier is None else peltier * -0.001830011 + 51.27039922,
        "EB Internal Temp(°C)": converted(thermistor, "EB_INTERNAL_TRP_TEMP"),
        "PSU board Temp (°C)": converted(thermistor, "EB_PSU_BOARD_TEMP"),
        "EB +12V (V)": None,
        "EB -12V (V)": None,
        "EB 5V": None,
        " EB+3V3 (V)": None,
        "TEC Rail(V)": None,
        "DIG Temp (°C)": converted(ob_thermistor, "OB_DIGITAL_TRP"),
        "DETEC Temp (°C)": converted(ob_thermistor, "OB_DETECTOR_TRP"),
        "MECH Temp (°C)": converted(ob_thermistor, "OB_MECHANISM_TRP"),
        "MTR Temp (°C)": converted(ob_thermistor, "OB_MOTOR_TRP"),
        "OB +3V3 (V)": None,
        "OB +1V5 (V)": None,
    }
    scales = {
        "EB +12V (V)": ("EB_MEAS_MAIN_12V", 0.000400543),
        "EB -12V (V)": ("EB_MEAS_MAIN_NEG12V", -0.00038147),
        "EB 5V": ("EB_MEAS_5V", 0.000152829),
        " EB+3V3 (V)": ("EB_MEAS_3V3", 0.0000763),
        "TEC Rail(V)": ("EB_MEAS_TEC_RAIL", 0.0000763),
        "OB +3V3 (V)": ("OB_3V3_VOLTAGE", 0.002),
        "OB +1V5 (V)": ("OB_1V5_VOLTAGE", 0.001),
    }
    for label, (name, scale) in scales.items():
        raw = attr(name)
        result[label] = None if raw is None else raw * scale
    return result


def _iter_numeric_fields(value: Any, depth: int = 0) -> Iterable[tuple[str, float]]:
    if depth > 3 or value is None:
        return
    dtype_names = getattr(getattr(value, "dtype", None), "names", None)
    if dtype_names:
        # NumPy structured scalars/arrays used by the eb_sci decoder.
        if getattr(value, "shape", ()) not in ((), (1,)):
            try:
                for record in value.flat:
                    yield from _iter_numeric_fields(record, depth + 1)
            except (AttributeError, TypeError):
                pass
            return
        for name in dtype_names:
            try:
                child = value[name]
                if getattr(child, "size", None) == 1:
                    child = child.item()
            except (IndexError, KeyError, TypeError, ValueError):
                continue
            number = finite(child)
            if number is not None:
                yield str(name), number
            elif not isinstance(child, (str, bytes, bytearray)):
                yield from _iter_numeric_fields(child, depth + 1)
        return
    if isinstance(value, dict):
        items = value.items()
    elif isinstance(value, (list, tuple)):
        for item in value:
            yield from _iter_numeric_fields(item, depth + 1)
        return
    elif callable(getattr(value, "_asdict", None)):
        items = value._asdict().items()
    else:
        try:
            items = vars(value).items()
        except TypeError:
            return
    for name, child in items:
        number = finite(child)
        if number is not None:
            yield str(name), number
        elif not isinstance(child, (str, bytes, bytearray)):
            yield from _iter_numeric_fields(child, depth + 1)


def science_raw_temperatures(*science_objects: Any) -> dict[str, float | None]:
    fields: list[tuple[str, str, float]] = []
    for science in science_objects:
        for name, value in _iter_numeric_fields(science):
            normalized = re.sub(r"[^A-Z0-9]", "", name.upper())
            fields.append((name, normalized, value))

    result: dict[str, float | None] = {}
    for label, candidates in SCIENCE_TEMP_FIELDS.items():
        normalized_candidates = {re.sub(r"[^A-Z0-9]", "", candidate.upper()) for candidate in candidates}
        value = next(
            (field_value for _name, normalized, field_value in fields if normalized in normalized_candidates),
            None,
        )
        result[label] = value
    return result


def science_temperatures_from_payload(payload: bytes) -> dict[str, float | None]:
    """Read the six >u16 eb_sci temperatures directly from fixed byte offsets."""
    offsets = {
        "HTSNK_START": 38,
        "HTSNK_END": 40,
        "SWIR_START": 42,
        "SWIR_END": 44,
        "MWIR_START": 46,
        "MWIR_END": 48,
    }
    result: dict[str, float | None] = {}
    for name, offset in offsets.items():
        result[name] = (
            float(int.from_bytes(payload[offset : offset + 2], byteorder="big", signed=False))
            if len(payload) >= offset + 2
            else None
        )
    return result


def summarise(path: Path, decoder: Any, mode: str) -> Summary:
    hk_packets: list[Packet] = []
    science_count = 0
    science_temps: list[dict[str, float | None]] = []
    warnings: list[str] = []
    for timestamp, payload, packet_type in payloads_from_log(path):
        try:
            if packet_type in (0x1, 0x2):
                hk_packets.append(Packet(timestamp, decoder.parse_eb_hk(payload)))
            elif packet_type in (0x4, 0x5, 0x6):
                science_count += 1
                packet_temperatures = {name: None for name in SCIENCE_OUTPUT_HEADERS}
                fn = {
                    0x4: getattr(decoder, "decode_dump_data", None),
                    0x5: getattr(decoder, "decode_cscience_data", None),
                    0x6: getattr(decoder, "decode_ncscience_data", None),
                }[packet_type]
                if callable(fn):
                    try:
                        decoded = fn(payload)
                        merge = getattr(decoder, "merge_sci_data_packet", None)
                        science = merge(decoded) if callable(merge) else decoded
                        decoded_temperatures = science_raw_temperatures(science, decoded)
                        for name, value in decoded_temperatures.items():
                            if value is not None:
                                packet_temperatures[name] = value
                    except Exception as exc:
                        if len(warnings) < 3:
                            warnings.append(f"science decode: {type(exc).__name__}")
                # The project decoder/merge path is authoritative. Retain the
                # fixed eb_sci layout only as a fallback for an older utility
                # module that does not expose all six header fields.
                raw_temperatures = science_temperatures_from_payload(payload)
                for name, value in raw_temperatures.items():
                    if packet_temperatures[name] is None:
                        packet_temperatures[name] = value
                science_temps.append(packet_temperatures)
        except Exception as exc:
            if len(warnings) < 3:
                warnings.append(f"decode: {type(exc).__name__}")

    rows = [convert_hk(packet.hk, decoder) for packet in hk_packets]
    ob3 = [row.get("OB +3V3 (V)") for row in rows]
    ob15 = [row.get("OB +1V5 (V)") for row in rows]
    switch_index = next(
        (i for i, value in enumerate(ob3) if finite(value) is not None and float(value) > 1.0),
        None,
    )
    ob_switch_time = (
        hk_packets[switch_index].time if switch_index is not None and switch_index < len(hk_packets) else None
    )

    values: dict[str, Any] = {}
    measurement_labels = [label for label in HEADERS if label in (rows[0] if rows else {})]
    for label in measurement_labels:
        if label in OB_FIRST_TEMPERATURE_HEADERS:
            # Ignore invalid/unpowered OB readings and take the first valid
            # value at or after the first powered OB housekeeping packet.
            ob_rows = rows[switch_index:] if switch_index is not None else []
            values[label] = choose((row.get(label) for row in ob_rows), "first")
        else:
            values[label] = choose((row.get(label) for row in rows), mode)
    if switch_index is None:
        warnings.append("OB switch transition not found")

    values["OB +3V3 Max(V)"] = max((n for v in ob3 if (n := finite(v)) is not None), default=None)
    values["OB +1V5 Max (V)"] = max((n for v in ob15 if (n := finite(v)) is not None), default=None)
    values["+3V3 Max Delta"] = (
        None
        if values.get("OB +3V3 Max(V)") is None or values.get("OB +3V3 (V)") is None
        else values["OB +3V3 Max(V)"] - values["OB +3V3 (V)"]
    )
    values["+1V5 Max Delta"] = (
        None
        if values.get("OB +1V5 Max (V)") is None or values.get("OB +1V5 (V)") is None
        else values["OB +1V5 Max (V)"] - values["OB +1V5 (V)"]
    )
    values["Sci Packets"] = science_count
    # All six photodiode temperatures must describe one acquisition. Use the
    # first science packet that exposes any of these fields; do not combine
    # start/end values from different packets.
    first_science_temperatures = next(
        (
            item
            for item in science_temps
            if any(item.get(field_name) is not None for field_name in SCIENCE_OUTPUT_HEADERS)
        ),
        None,
    )
    for field_name, output_header in SCIENCE_OUTPUT_HEADERS.items():
        values[output_header] = (
            first_science_temperatures.get(field_name) if first_science_temperatures is not None else None
        )
    if science_count and not any(
        values.get(output_header) is not None for output_header in SCIENCE_OUTPUT_HEADERS.values()
    ):
        warnings.append("science packets decoded but photodiode temperature fields were not found")

    times = [packet.time for packet in hk_packets if packet.time is not None]
    if not hk_packets:
        warnings.append("no HK packets")
    return Summary(
        values,
        len(hk_packets),
        science_count,
        min(times, default=None),
        max(times, default=None),
        ob_switch_time,
        warnings,
        detect_state_condition_times(hk_packets),
    )


def _parse_msc_csv(path: Path, value_column: int) -> list[TemperaturePoint]:
    """Read an MSC CSV using a zero-based temperature column index."""
    points: list[TemperaturePoint] = []
    with path.open(newline="", encoding="utf-8-sig", errors="ignore") as handle:
        for row in csv.reader(handle):
            if len(row) <= value_column:
                continue
            timestamp_text = row[0].strip()
            timestamp = None
            for fmt in ("%Y%m%d_%H%M%S", "%Y-%m-%d %H:%M:%S"):
                try:
                    timestamp = datetime.strptime(timestamp_text, fmt)
                    break
                except ValueError:
                    continue
            if timestamp is None:
                continue
            value = finite(row[value_column])
            if value is None or value <= -150.0 or value >= 150.0:
                continue
            points.append(TemperaturePoint(timestamp, value, path))
    return points


def _tp1000_date(path: Path) -> datetime | None:
    match = re.search(r"(\d{2})-(\d{2})-(\d{2})", path.name)
    if match:
        day, month, year = (int(part) for part in match.groups())
        return datetime(2000 + year, month, day)
    for part in reversed(path.parts[:-1]):
        match = re.fullmatch(r"(\d{2})[_-](\d{2})[_-](\d{4})", part)
        if match:
            day, month, year = (int(value) for value in match.groups())
            return datetime(year, month, day)
    return None


def _parse_tp1000(path: Path) -> list[TemperaturePoint]:
    """Read TP1000 text or CSV rows."""
    base_date = _tp1000_date(path)
    points: list[TemperaturePoint] = []
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        parts = re.split(r"[\t,; ]+", line.strip())
        if len(parts) < 2:
            continue
        full_timestamp = " ".join(parts[:2])
        parsed_full = None
        for text in (parts[0], full_timestamp):
            for fmt in ("%Y%m%d_%H%M%S", "%Y-%m-%d %H:%M:%S"):
                try:
                    parsed_full = datetime.strptime(text, fmt)
                    break
                except ValueError:
                    continue
            if parsed_full is not None:
                break
        if parsed_full is not None:
            value_index = 1 if "_" in parts[0] else 2
            value = finite(parts[value_index]) if len(parts) > value_index else None
            if value is not None and -150.0 < value < 150.0:
                points.append(TemperaturePoint(parsed_full, value, path))
            continue

        match = re.fullmatch(r"(\d{2})[.:](\d{2})[.:](\d{2})", parts[0])
        value = finite(parts[1])
        if base_date is None or not match or value is None or value <= -150.0 or value >= 150.0:
            continue
        hour, minute, second = (int(part) for part in match.groups())
        timestamp = base_date + timedelta(hours=hour, minutes=minute, seconds=second)
        points.append(TemperaturePoint(timestamp, value, path))
    return points


def _spreadsheet_date(value: Any) -> date | None:
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    if value is None:
        return None
    text = str(value).strip()
    if re.fullmatch(r"\d+(?:\.0+)?", text):
        text = str(int(float(text))).zfill(8)
    for fmt in ("%d%m%Y", "%Y%m%d", "%d-%m-%Y", "%d/%m/%Y", "%Y-%m-%d"):
        try:
            return datetime.strptime(text, fmt).date()
        except ValueError:
            continue
    return None


def _spreadsheet_time(value: Any) -> timedelta | None:
    if isinstance(value, datetime):
        value = value.time()
    if isinstance(value, time):
        return timedelta(
            hours=value.hour,
            minutes=value.minute,
            seconds=value.second,
            microseconds=value.microsecond,
        )
    if isinstance(value, (int, float)) and math.isfinite(float(value)):
        fraction = float(value) % 1.0
        return timedelta(seconds=round(fraction * 24 * 60 * 60))
    if value is None:
        return None
    text = str(value).strip()
    try:
        numeric = float(text)
    except ValueError:
        numeric = None
    if numeric is not None and math.isfinite(numeric):
        return timedelta(seconds=round((numeric % 1.0) * 24 * 60 * 60))
    match = re.fullmatch(r"(\d{1,2})[.:](\d{2})[.:](\d{2})(?:[.,](\d{1,6}))?", text)
    if not match:
        return None
    hour, minute, second = (int(match.group(index)) for index in (1, 2, 3))
    micros = int((match.group(4) or "").ljust(6, "0") or 0)
    if hour > 23 or minute > 59 or second > 59:
        return None
    return timedelta(hours=hour, minutes=minute, seconds=second, microseconds=micros)


def _spreadsheet_timestamp(
    date_value: Any,
    time_value: Any,
    fallback_date: date | None,
) -> datetime | None:
    if isinstance(date_value, datetime) and time_value in (None, ""):
        return date_value
    day = _spreadsheet_date(date_value) or fallback_date
    clock = _spreadsheet_time(time_value)
    if day is None or clock is None:
        return None
    return datetime.combine(day, time()) + clock


def _date_from_sheet_name(name: str) -> date | None:
    match = re.search(r"(?<!\d)(\d{2})[-_](\d{2})[-_](\d{2}|\d{4})(?!\d)", name)
    if not match:
        return None
    day, month, year = (int(value) for value in match.groups())
    if year < 100:
        year += 2000
    try:
        return date(year, month, day)
    except ValueError:
        return None


def _workbook_table_dates(path: Path) -> dict[tuple[str, int], date]:
    """Map (worksheet name, table header row) to dates embedded in table names."""
    main_ns = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    office_rel_ns = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    package_rel_ns = "http://schemas.openxmlformats.org/package/2006/relationships"
    result: dict[tuple[str, int], date] = {}
    try:
        with zipfile.ZipFile(path) as archive:
            workbook_root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
            workbook_rels = ElementTree.fromstring(archive.read("xl/_rels/workbook.xml.rels"))
            workbook_targets = {
                relation.attrib["Id"]: relation.attrib["Target"]
                for relation in workbook_rels.findall(f"{{{package_rel_ns}}}Relationship")
            }
            sheets = workbook_root.find(f"{{{main_ns}}}sheets")
            if sheets is None:
                return result
            for sheet in sheets:
                sheet_name = sheet.attrib.get("name", "")
                relation_id = sheet.attrib.get(f"{{{office_rel_ns}}}id")
                target = workbook_targets.get(relation_id or "")
                if not target or "worksheets/" not in target:
                    continue
                sheet_path = posixpath.normpath(posixpath.join("xl", target.lstrip("/")))
                rels_path = posixpath.join(
                    posixpath.dirname(sheet_path),
                    "_rels",
                    posixpath.basename(sheet_path) + ".rels",
                )
                if rels_path not in archive.namelist():
                    continue
                sheet_root = ElementTree.fromstring(archive.read(sheet_path))
                sheet_rels = ElementTree.fromstring(archive.read(rels_path))
                sheet_targets = {
                    relation.attrib["Id"]: relation.attrib["Target"]
                    for relation in sheet_rels.findall(f"{{{package_rel_ns}}}Relationship")
                }
                for table_part in sheet_root.findall(f".//{{{main_ns}}}tablePart"):
                    table_relation = table_part.attrib.get(f"{{{office_rel_ns}}}id")
                    table_target = sheet_targets.get(table_relation or "")
                    if not table_target:
                        continue
                    table_path = posixpath.normpath(posixpath.join(posixpath.dirname(sheet_path), table_target))
                    table_root = ElementTree.fromstring(archive.read(table_path))
                    table_date = _date_from_sheet_name(table_root.attrib.get("displayName", ""))
                    reference = table_root.attrib.get("ref", "")
                    row_match = re.search(r"\d+", reference)
                    if table_date is not None and row_match:
                        result[(sheet_name, int(row_match.group()))] = table_date
    except (KeyError, OSError, zipfile.BadZipFile, ElementTree.ParseError):
        return result
    return result


def _parse_temperature_workbook(
    path: Path,
    *,
    chamber_source: str | None,
) -> dict[str, list[TemperaturePoint]]:
    """Read full EB/OB logger workbooks and embedded TP1000 worksheets."""
    parsed: dict[str, list[TemperaturePoint]] = {
        "eb_chamber_temp": [],
        "ob_chamber_temp": [],
        "rov_therm": [],
    }
    table_dates = _workbook_table_dates(path)
    warning_context = warnings.catch_warnings()
    warning_context.__enter__()
    warnings.simplefilter("ignore", UserWarning)
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        for worksheet in workbook.worksheets:
            sheet_date = _date_from_sheet_name(worksheet.title)
            last_date = sheet_date
            rov_time_column: int | None = None
            rov_temp_column: int | None = None
            rov_date = sheet_date

            for row_number, row in enumerate(
                worksheet.iter_rows(values_only=True),
                start=1,
            ):
                values = list(row)
                if not values:
                    continue

                row_date = _spreadsheet_date(values[0])
                if row_date is not None:
                    last_date = row_date

                normalized_headers = [
                    re.sub(r"[^a-z0-9]", "", str(value).lower()) if isinstance(value, str) else "" for value in values
                ]
                temperature_columns = [
                    index
                    for index, heading in enumerate(normalized_headers)
                    if heading in ("temperature", "tmperature")
                ]
                timestamp_columns = [
                    index for index, heading in enumerate(normalized_headers) if heading == "timestamp"
                ]
                if temperature_columns and timestamp_columns:
                    rov_temp_column = temperature_columns[0]
                    # Some imported TP1000 tables contain an extra Column1 and
                    # chart the column immediately before Tmperature.
                    rov_time_column = (
                        rov_temp_column - 1 if rov_temp_column - 1 > timestamp_columns[0] else timestamp_columns[0]
                    )
                    rov_date = table_dates.get((worksheet.title, row_number)) or sheet_date or last_date
                    continue

                if chamber_source is not None and len(values) > 6:
                    timestamp = _spreadsheet_timestamp(
                        values[0],
                        values[1] if len(values) > 1 else None,
                        last_date,
                    )
                    chamber_value = finite(values[6])  # Excel column G
                    if timestamp is not None and chamber_value is not None and -150.0 < chamber_value < 150.0:
                        parsed[chamber_source].append(TemperaturePoint(timestamp, chamber_value, path))

                if (
                    rov_time_column is not None
                    and rov_temp_column is not None
                    and len(values) > max(rov_time_column, rov_temp_column)
                ):
                    timestamp = _spreadsheet_timestamp(
                        None,
                        values[rov_time_column],
                        rov_date or last_date,
                    )
                    rov_value = finite(values[rov_temp_column])
                    if timestamp is not None and rov_value is not None and -150.0 < rov_value < 150.0:
                        parsed["rov_therm"].append(TemperaturePoint(timestamp, rov_value, path))
    finally:
        if workbook is not None:
            workbook.close()
        warning_context.__exit__(None, None, None)
    return parsed


def discover_temperature_metadata(
    root: Path,
) -> dict[str, list[TemperaturePoint]]:
    """Discover CSV/TXT and full XLSX EB, OB, and ROV temperature logs."""
    result: dict[str, list[TemperaturePoint]] = {
        "eb_chamber_temp": [],
        "ob_chamber_temp": [],
        "rov_therm": [],
    }
    workbook_sources: dict[Path, str | None] = {}
    for path in root.rglob("*"):
        if not path.is_file():
            continue
        if path.name.startswith("~$"):
            continue
        if re.search(r"EM.?EQM.*THERMAL.*SUMMARY", path.stem, re.IGNORECASE):
            continue
        folder_names = {part.upper() for part in path.parts}
        normalized_folders = {re.sub(r"[^A-Z0-9]", "", name) for name in folder_names}
        in_msc2 = any(name.startswith("MSC2") for name in normalized_folders)
        in_msc1 = any(name.startswith("MSC1") for name in normalized_folders)
        in_tp1000 = any("TP1000" in name or name == "TRPOB" for name in normalized_folders)
        if in_msc2 and path.suffix.lower() == ".csv":
            # EB chamber temperature is recorded in Excel column G.
            result["eb_chamber_temp"].extend(_parse_msc_csv(path, value_column=6))
        elif in_msc1 and path.suffix.lower() == ".csv":
            result["ob_chamber_temp"].extend(_parse_msc_csv(path, value_column=1))
        elif in_tp1000 and path.suffix.lower() in (".txt", ".csv"):
            result["rov_therm"].extend(_parse_tp1000(path))
        elif path.suffix.lower() in (".xlsx", ".xlsm"):
            normalized_name = re.sub(r"[^A-Z0-9]", "", path.stem.upper())
            if in_msc2 or re.search(r"(^|[^A-Z])EB([^A-Z]|$)", path.stem.upper()):
                chamber_source = "eb_chamber_temp"
            elif in_msc1 or re.search(r"(^|[^A-Z])OB([^A-Z]|$)", path.stem.upper()):
                chamber_source = "ob_chamber_temp"
            elif normalized_name.startswith("EB"):
                chamber_source = "eb_chamber_temp"
            elif normalized_name.startswith("OB"):
                chamber_source = "ob_chamber_temp"
            else:
                chamber_source = None
            workbook_sources[path.resolve()] = chamber_source
    for workbook_path, chamber_source in sorted(workbook_sources.items(), key=lambda item: str(item[0]).lower()):
        print(f"Reading temperature workbook: {workbook_path.name}")
        try:
            workbook_data = _parse_temperature_workbook(
                workbook_path,
                chamber_source=chamber_source,
            )
        except Exception as exc:
            print(f"  skipped {workbook_path.name}: {type(exc).__name__}: {exc}")
            continue
        print(
            "  samples: "
            f"EB={len(workbook_data['eb_chamber_temp'])}, "
            f"OB={len(workbook_data['ob_chamber_temp'])}, "
            f"ROV={len(workbook_data['rov_therm'])}"
        )
        for key, points in workbook_data.items():
            result[key].extend(points)
    for points in result.values():
        points.sort(key=lambda point: point.time)
    return result


def _metadata_value_for_run(
    points: list[TemperaturePoint],
    start: datetime,
    end: datetime,
    *,
    match_by_date: bool = False,
    fallback_by_date: bool = False,
    nearest_hours: float = 2.0,
) -> float | None:
    if not points:
        return None
    if match_by_date:
        values = [point.value for point in points if point.time.date() == start.date()]
        return choose(values, "median")

    values = [point.value for point in points if start <= point.time <= end]
    if values:
        return choose(values, "median")
    midpoint = start + (end - start) / 2
    if fallback_by_date:
        same_day = [point for point in points if point.time.date() == start.date()]
        if same_day:
            nearest_same_day = min(
                same_day,
                key=lambda point: abs((point.time - midpoint).total_seconds()),
            )
            return nearest_same_day.value
    nearest = min(points, key=lambda point: abs((point.time - midpoint).total_seconds()))
    if abs((nearest.time - midpoint).total_seconds()) <= nearest_hours * 60 * 60:
        return nearest.value
    return None


def _derive_chamber_setpoint(*values: Any) -> float | None:
    readings = [number for value in values if (number := finite(value)) is not None]
    if not readings:
        return None
    representative = statistics.median(readings)
    if 20.0 <= representative <= 30.0:
        return 25.0
    return float(round(representative / 10.0) * 10)


def _setpoint_from_run_name(run_name: str) -> float | None:
    text = run_name.strip()
    if re.search(r"\bambient\b", text, re.IGNORECASE):
        return 25.0
    # Remove the SFT identifier (including ranges such as SFT3-4), then use
    # the first stated chamber condition. This preserves the initial condition
    # for delta runs such as "-40oC +20oC".
    conditions = re.sub(r"^\s*SFT\s*\d+(?:-\d+)?", "", text, flags=re.IGNORECASE)
    match = re.search(r"(?<!\d)([-+]?\s*\d+(?:\.\d+)?)\s*(?:°|o)?\s*C?\b", conditions, re.IGNORECASE)
    if not match:
        return None
    return finite(match.group(1).replace(" ", ""))


def _last_temperature_before(
    points: list[TemperaturePoint],
    target: datetime,
    *,
    max_age_hours: float = 24.0,
) -> float | None:
    eligible = [point for point in points if point.time < target]
    if not eligible:
        return None
    latest = max(eligible, key=lambda point: point.time)
    if (target - latest.time).total_seconds() > max_age_hours * 60 * 60:
        return None
    return latest.value


def _rov_temperature_and_dwell_before(
    points: list[TemperaturePoint],
    target: datetime,
    *,
    tolerance_c: float = 1.0,
    max_gap_s: float = 120.0,
) -> tuple[float | None, float | None]:
    eligible = sorted(
        (point for point in points if point.time < target and point.time.date() == target.date()),
        key=lambda point: point.time,
    )
    if not eligible:
        return None, None
    latest = eligible[-1]
    if (target - latest.time).total_seconds() > 24 * 60 * 60:
        return None, None
    # The dwell is only demonstrable while the TP1000 samples remain
    # continuous up to the heater event. A stale last sample can still supply
    # the closest pre-event temperature, but cannot prove a dwell duration.
    latest_gap_s = (target - latest.time).total_seconds()
    if latest_gap_s > max_gap_s:
        return latest.value, None

    dwell_start = latest.time
    next_time = latest.time
    for point in reversed(eligible[:-1]):
        if (target - point.time).total_seconds() > 60 * 60:
            break
        gap_s = (next_time - point.time).total_seconds()
        if gap_s > max_gap_s or abs(point.value - latest.value) > tolerance_c:
            break
        dwell_start = point.time
        next_time = point.time
    dwell_minutes = max(0.0, (target - dwell_start).total_seconds() / 60.0)
    return latest.value, dwell_minutes


def metadata_for(
    log: Path,
    result: Summary,
    metadata: dict[str, list[TemperaturePoint]],
    temperature_reference_time: datetime | None,
) -> dict[str, Any]:
    start = result.first_time or parse_filename_time(log)
    end = result.last_time or start
    output: dict[str, Any] = {"run": log.parent.name or log.stem}
    if start is None or end is None:
        return output
    output["SFT Date/Time"] = start
    if temperature_reference_time is not None:
        output["EB Chamber Temp (°C)"] = _last_temperature_before(
            metadata["eb_chamber_temp"], temperature_reference_time
        )
        output["OB Chamber Temp(°C)"] = _last_temperature_before(
            metadata["ob_chamber_temp"], temperature_reference_time
        )
        rov_temp, rov_dwell = _rov_temperature_and_dwell_before(metadata["rov_therm"], temperature_reference_time)
        output["ROV_THERM (°C)"] = rov_temp
        output["ROV Dwell Before CH3/CH4 (min)"] = rov_dwell
        output["ROV Dwell >10 min"] = None if rov_dwell is None else rov_dwell > 10.0
    else:
        output["EB Chamber Temp (°C)"] = None
        output["OB Chamber Temp(°C)"] = None
        output["ROV_THERM (°C)"] = None
        output["ROV Dwell Before CH3/CH4 (min)"] = None
        output["ROV Dwell >10 min"] = None
    output["Temp (°C)"] = _setpoint_from_run_name(output["run"])
    if output["Temp (°C)"] is None:
        output["Temp (°C)"] = _derive_chamber_setpoint(
            output.get("EB Chamber Temp (°C)"),
            output.get("OB Chamber Temp(°C)"),
            output.get("ROV_THERM (°C)"),
        )
    return output


def _local_psu_logs(root: Path, log: Path) -> list[Path]:
    candidates = [
        path for path in root.rglob("*") if path.is_file() and re.search(r"_PSU\.log$", path.name, re.IGNORECASE)
    ]

    def locality(path: Path) -> tuple[int, int, str]:
        if path.parent == log.parent:
            rank = 0
        elif path.parent.parent == log.parent or log.parent.parent == path.parent:
            rank = 1
        elif path.parent.parent == log.parent.parent:
            rank = 2
        else:
            rank = 3
        return rank, len(path.parts), str(path).lower()

    return sorted(candidates, key=locality)


def _read_psu_channel_samples(path: Path, channel: str) -> list[tuple[datetime, float, float]]:
    samples: list[tuple[datetime, float, float]] = []
    timestamp_pattern = r"(?P<ts>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2},\d+)"
    number = r"[-+]?\d*\.?\d+"
    compact = re.compile(
        rf"^{timestamp_pattern}\s+-\s+{channel}\s+"
        rf"(?P<v>{number})(?:V)?\s+(?P<i>{number})(?:A)?",
        re.IGNORECASE,
    )
    labelled = re.compile(
        rf"^{timestamp_pattern}\s+-\s+{channel}\s+Voltage:\s*(?P<v>{number})(?:V)?\s+"
        rf"{channel}\s+Current:\s*(?P<i>{number})(?:A)?",
        re.IGNORECASE,
    )
    dual = re.compile(
        rf"^{timestamp_pattern}\s+-\s+"
        rf"CH3\s+Voltage:\s*(?P<ch3v>{number})(?:V)?\s+"
        rf"CH3\s+Current:\s*(?P<ch3i>{number})(?:A)?\s+"
        rf"CH4\s+Voltage:\s*(?P<ch4v>{number})(?:V)?\s+"
        rf"CH4\s+Current:\s*(?P<ch4i>{number})(?:A)?",
        re.IGNORECASE,
    )
    for line in path.read_text(encoding="utf-8", errors="ignore").splitlines():
        stripped = line.strip()
        match = dual.search(stripped)
        try:
            if match:
                voltage = float(match.group("ch3v" if channel == "CH3" else "ch4v"))
                current = float(match.group("ch3i" if channel == "CH3" else "ch4i"))
            else:
                match = labelled.search(stripped) or compact.search(stripped)
                if not match:
                    continue
                voltage = float(match.group("v"))
                current = float(match.group("i"))
            timestamp = datetime.strptime(match.group("ts"), "%Y-%m-%d %H:%M:%S,%f")
        except (ValueError, AttributeError):
            continue
        samples.append((timestamp, voltage, current))
    return samples


def find_channel_turn_on(
    root: Path,
    log: Path,
    start: datetime | None,
    end: datetime | None,
    channel: str,
) -> datetime | None:
    channel = channel.upper()
    if channel not in {"CH3", "CH4"}:
        raise ValueError(f"Unsupported PSU channel: {channel}")
    if start is None:
        start = parse_filename_time(log)
    if start is None:
        return None
    end = end or start + timedelta(hours=1)
    window_start = start - timedelta(hours=1)
    window_end = end + timedelta(minutes=30)
    candidates: list[tuple[int, float, datetime]] = []
    for locality_rank, psu_log in enumerate(_local_psu_logs(root, log)):
        samples = sorted(
            (
                sample
                for sample in _read_psu_channel_samples(psu_log, channel)
                if window_start <= sample[0] <= window_end
            ),
            key=lambda sample: sample[0],
        )
        if not samples:
            continue
        previous_powered: bool | None = None
        transitions: list[datetime] = []
        for timestamp, voltage, current in samples:
            powered = voltage > 1.0 or current > 0.001
            if powered and previous_powered is False:
                transitions.append(timestamp)
            previous_powered = powered
        for transition in transitions:
            # Prefer a transition before/inside the RS422 run, then the nearest
            # local PSU file and smallest time distance.
            timing_rank = 0 if transition <= end else 1
            distance = abs((transition - start).total_seconds())
            candidates.append((timing_rank * 1000 + locality_rank, distance, transition))
    return min(candidates, default=(0, 0.0, None), key=lambda item: (item[0], item[1]))[2]


def nearest_psu_current(root: Path, log: Path, target: datetime | None) -> float | None:
    if target is None:
        return None
    candidates = _local_psu_logs(root, log)
    compact = re.compile(
        r"^(?P<ts>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2},\d+)\s+-\s+"
        r"CH4\s+[-+]?\d*\.?\d+(?:V)?\s+(?P<i>[-+]?\d*\.?\d+)(?:A)?",
        re.IGNORECASE,
    )
    labelled = re.compile(
        r"^(?P<ts>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2},\d+)\s+-\s+"
        r"CH4\s+Voltage:\s*[-+]?\d*\.?\d+(?:V)?\s+"
        r"CH4\s+Current:\s*(?P<i>[-+]?\d*\.?\d+)(?:A)?",
        re.IGNORECASE,
    )
    dual = re.compile(
        r"^(?P<ts>\d{4}-\d{2}-\d{2} \d{2}:\d{2}:\d{2},\d+)\s+-\s+"
        r"CH3\s+Voltage:\s*[-+]?\d*\.?\d+(?:V)?\s+"
        r"CH3\s+Current:\s*[-+]?\d*\.?\d+(?:A)?\s+"
        r"CH4\s+Voltage:\s*[-+]?\d*\.?\d+(?:V)?\s+"
        r"CH4\s+Current:\s*(?P<i>[-+]?\d*\.?\d+)(?:A)?",
        re.IGNORECASE,
    )
    best: tuple[float, float] | None = None
    for candidate in candidates[:20]:
        for line in candidate.read_text(encoding="utf-8", errors="ignore").splitlines():
            stripped = line.strip()
            match = dual.search(stripped) or labelled.search(stripped) or compact.search(stripped)
            if not match:
                continue
            try:
                timestamp = datetime.strptime(match.group("ts"), "%Y-%m-%d %H:%M:%S,%f")
                current = float(match.group("i"))
            except ValueError:
                continue
            distance = abs((timestamp - target).total_seconds())
            if best is None or distance < best[0]:
                best = (distance, current)
    return best[1] if best and best[0] <= 300 else None


def _state_current_from_samples(
    samples: list[tuple[datetime, float, float]],
    target: datetime,
    *,
    duration_s: float = FIXED_STATE_CURRENT_WINDOW_S,
) -> float | None:
    """Return CH4/EB PSU current in mA over the post-condition check window."""
    window_end = target + timedelta(seconds=duration_s)
    window_currents = [current for timestamp, _voltage, current in samples if target <= timestamp <= window_end]
    if window_currents:
        return statistics.fmean(window_currents) * 1000.0
    nearest = min(
        ((abs((timestamp - target).total_seconds()), current) for timestamp, _voltage, current in samples),
        default=None,
        key=lambda item: item[0],
    )
    return nearest[1] * 1000.0 if nearest and nearest[0] <= duration_s else None


def _state_current_at_condition(
    root: Path,
    log: Path,
    target: datetime,
    *,
    duration_s: float = FIXED_STATE_CURRENT_WINDOW_S,
) -> float | None:
    samples = sorted(
        {
            sample
            for candidate in _local_psu_logs(root, log)[:20]
            for sample in _read_psu_channel_samples(candidate, "CH4")
        },
        key=lambda sample: sample[0],
    )
    return _state_current_from_samples(samples, target, duration_s=duration_s)


def offline_state_current_metadata(
    root: Path,
    log: Path,
    summary: Summary,
) -> dict[str, float | None]:
    """Perform state-current extraction from detected HK conditions plus PSU logs."""
    samples = sorted(
        {
            sample
            for candidate in _local_psu_logs(root, log)[:20]
            for sample in _read_psu_channel_samples(candidate, "CH4")
        },
        key=lambda sample: sample[0],
    )
    output: dict[str, float | None] = {}
    for label in STATE_CHECK_LABELS:
        condition_time = summary.state_condition_times.get(label)
        output[f"{label} Current (mA)"] = (
            _state_current_from_samples(samples, condition_time) if condition_time is not None else None
        )
    return output


def _sft_number(row: tuple[Path, Summary, dict[str, Any]]) -> int | None:
    log, _summary, metadata = row
    run_name = str(metadata.get("run") or log.parent.name or log.stem)
    match = re.search(r"\bSFT\s*(\d+)", run_name, re.IGNORECASE)
    return int(match.group(1)) if match else None


def _row_datetime(row: tuple[Path, Summary, dict[str, Any]]) -> datetime:
    log, summary, metadata = row
    value = metadata.get("SFT Date/Time") or summary.first_time or parse_filename_time(log)
    return value if isinstance(value, datetime) else datetime.min


def group_rows_by_test_matrix(
    rows: list[tuple[Path, Summary, dict[str, Any]]],
) -> list[tuple[Path, Summary, dict[str, Any]]]:
    """Select the preferred SFT per matrix cell and backfill its blank fields."""
    rows_by_sft: dict[int, list[tuple[Path, Summary, dict[str, Any]]]] = {}
    for row in rows:
        number = _sft_number(row)
        if number is not None and number != 10:
            rows_by_sft.setdefault(number, []).append(row)
    for sft_rows in rows_by_sft.values():
        sft_rows.sort(key=_row_datetime, reverse=True)

    grouped: list[tuple[Path, Summary, dict[str, Any]]] = []
    identity_fields = {
        "run",
        "SFT Date/Time",
        "Temp (°C)",
        "Matrix EB Setpoint (°C)",
        "Matrix OB Setpoint (°C)",
        "Backfilled From",
    }

    for eb_setpoint, ob_setpoint, preferences in TEST_MATRIX_GROUPS:
        available = [(sft, rows_by_sft[sft][0]) for sft in preferences if rows_by_sft.get(sft)]
        if not available:
            continue

        primary_sft, primary_row = available[0]
        log, summary, metadata = primary_row
        combined = {**summary.values, **metadata}
        backfill_notes: list[str] = []
        if primary_sft != preferences[0]:
            backfill_notes.append(f"SFT{preferences[0]} unavailable; used SFT{primary_sft}")

        for backup_sft, backup_row in available[1:]:
            _backup_log, backup_summary, backup_metadata = backup_row
            backup = {**backup_summary.values, **backup_metadata}
            filled_fields: list[str] = []
            for heading in HEADERS:
                if heading in identity_fields:
                    continue
                if combined.get(heading) in (None, "") and backup.get(heading) not in (None, ""):
                    combined[heading] = backup[heading]
                    filled_fields.append(heading)
            if filled_fields:
                backfill_notes.append(f"SFT{backup_sft}: " + ", ".join(filled_fields))

        combined["Temp (°C)"] = eb_setpoint
        combined["Matrix EB Setpoint (°C)"] = eb_setpoint
        combined["Matrix OB Setpoint (°C)"] = ob_setpoint
        combined["Backfilled From"] = "; ".join(backfill_notes)

        grouped_summary = Summary(
            values=combined,
            hk_count=summary.hk_count,
            science_packets=summary.science_packets,
            first_time=summary.first_time,
            last_time=summary.last_time,
            ob_switch_time=summary.ob_switch_time,
            warnings=summary.warnings,
            state_condition_times=summary.state_condition_times,
        )
        grouped.append((log, grouped_summary, {}))

    return grouped


def write_workbook(
    output: Path,
    rows: list[tuple[Path, Summary, dict[str, Any]]],
) -> None:
    """Create a standalone, formatted Summary workbook from the coded headers."""
    output.parent.mkdir(parents=True, exist_ok=True)
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.sheet_view.showGridLines = False
    summary_sheet.sheet_view.zoomScale = 85
    summary_sheet.append(HEADERS)

    for _log, summary, metadata in rows:
        merged = {**summary.values, **metadata}
        summary_sheet.append([merged.get(heading) for heading in HEADERS])

    navy = "1F4E78"
    pale_blue = "DCE6F1"
    light_border = Side(style="thin", color="D9E2F3")
    header_fill = PatternFill("solid", fgColor=navy)
    header_font = Font(name="Aptos", size=10, bold=True, color="FFFFFF")
    body_font = Font(name="Aptos", size=10, color="1F1F1F")
    summary_sheet.row_dimensions[1].height = 42

    for cell in summary_sheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = Border(bottom=Side(style="medium", color=navy))

    for row in summary_sheet.iter_rows(min_row=2, max_row=max(2, summary_sheet.max_row)):
        for cell in row:
            cell.font = body_font
            cell.alignment = Alignment(vertical="center")
            cell.border = Border(bottom=light_border)
        if row[0].row % 2 == 0:
            for cell in row:
                cell.fill = PatternFill("solid", fgColor="F7FAFC")

    width_overrides = {
        "Temp (°C)": 12,
        "run": 28,
        "SFT Date/Time": 21,
        "EB Chamber Temp (°C)": 22,
        "OB Chamber Temp(°C)": 22,
        "ROV_THERM (°C)": 18,
        "CH4_I at OB Switch on (A)": 25,
        "ROV Dwell Before CH3/CH4 (min)": 31,
        "ROV Dwell >10 min": 20,
        "Matrix EB Setpoint (°C)": 24,
        "Matrix OB Setpoint (°C)": 24,
        "Backfilled From": 70,
    }
    for column, heading in enumerate(HEADERS, start=1):
        if heading in STATE_CHECK_HEADERS:
            width = 34 if heading.startswith("State6 ") else 23
        elif heading in width_overrides:
            width = width_overrides[heading]
        elif "Temp" in heading or "THERM" in heading:
            width = 21
        elif "Voltage" in heading or " (V)" in heading or "Max" in heading:
            width = 19
        else:
            width = min(24, max(13, len(heading) + 2))
        summary_sheet.column_dimensions[get_column_letter(column)].width = width

    for row_index in range(2, summary_sheet.max_row + 1):
        for column, heading in enumerate(HEADERS, start=1):
            cell = summary_sheet.cell(row_index, column)
            if heading == "SFT Date/Time":
                cell.number_format = "yyyy-mm-dd hh:mm:ss"
            elif heading == "Sci Packets":
                cell.number_format = "0"
            elif heading == "ROV Dwell Before CH3/CH4 (min)":
                cell.number_format = "0.0"
            elif heading in STATE_CHECK_HEADERS:
                cell.number_format = "0.00"
            elif "Temp" in heading or "THERM" in heading or "Setpoint" in heading:
                cell.number_format = "0.00"
            elif "(A)" in heading or "(V)" in heading or "Delta" in heading:
                cell.number_format = "0.000"
            if heading == "Backfilled From":
                cell.alignment = Alignment(vertical="top", wrap_text=True)
            elif heading in ("run",):
                cell.alignment = Alignment(horizontal="left", vertical="center")
            elif heading == "ROV Dwell >10 min":
                cell.alignment = Alignment(horizontal="center", vertical="center")

    if rows:
        last_column = get_column_letter(len(HEADERS))
        table = Table(
            displayName="ThermalSummary",
            ref=f"A1:{last_column}{len(rows) + 1}",
        )
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        summary_sheet.add_table(table)

    summary_sheet.freeze_panes = "A2"
    summary_sheet.auto_filter.ref = f"A1:{get_column_letter(len(HEADERS))}{max(1, len(rows) + 1)}"
    summary_sheet.print_title_rows = "1:1"
    summary_sheet.page_setup.orientation = "landscape"
    # The summary is deliberately wide; four landscape pages preserve readable
    # print/PDF text instead of shrinking all 39 columns onto one page.
    summary_sheet.page_setup.fitToWidth = 4
    summary_sheet.page_setup.fitToHeight = 0
    summary_sheet.sheet_properties.pageSetUpPr.fitToPage = True

    workbook.save(output)


def select_inputs() -> Path | None:
    root = tk.Tk()
    root.title("RS422 Excel Extractor")
    root.withdraw()
    root.update_idletasks()
    try:
        root.attributes("-topmost", True)
    except tk.TclError:
        pass

    def ask_directory(title: str) -> str:
        print(f"Waiting for folder selection: {title}")
        root.update_idletasks()
        return filedialog.askdirectory(
            title=title,
            mustexist=True,
            parent=root,
        )

    try:
        logs_directory = ask_directory("Select top-level test folder (RS422, PSU, MSC1, MSC2 and TP1000 logs)")
        if not logs_directory:
            return None
        return Path(logs_directory)
    finally:
        root.destroy()


class ProgressDialog:
    """Small determinate progress window for the synchronous extraction flow."""

    def __init__(self) -> None:
        self.root = tk.Tk()
        self.root.title("RS422 Excel Extractor")
        self.root.resizable(False, False)
        self.root.protocol("WM_DELETE_WINDOW", lambda: None)
        try:
            self.root.attributes("-topmost", True)
        except tk.TclError:
            pass

        frame = ttk.Frame(self.root, padding=18)
        frame.grid(row=0, column=0, sticky="nsew")
        self.status = tk.StringVar(value="Starting extraction…")
        self.detail = tk.StringVar(value="")
        ttk.Label(frame, textvariable=self.status, font=("Segoe UI", 11, "bold")).grid(row=0, column=0, sticky="w")
        ttk.Label(
            frame,
            textvariable=self.detail,
            width=68,
            wraplength=520,
        ).grid(row=1, column=0, sticky="w", pady=(5, 12))
        self.bar = ttk.Progressbar(
            frame,
            orient="horizontal",
            mode="determinate",
            maximum=100.0,
            length=520,
        )
        self.bar.grid(row=2, column=0, sticky="ew")
        self.percent = tk.StringVar(value="0%")
        ttk.Label(frame, textvariable=self.percent).grid(row=3, column=0, sticky="e", pady=(5, 0))

        self.root.update_idletasks()
        width = self.root.winfo_reqwidth()
        height = self.root.winfo_reqheight()
        x = max(0, (self.root.winfo_screenwidth() - width) // 2)
        y = max(0, (self.root.winfo_screenheight() - height) // 2)
        self.root.geometry(f"{width}x{height}+{x}+{y}")
        self.root.update()

    def set(self, value: float, status: str, detail: str = "") -> None:
        value = min(100.0, max(0.0, float(value)))
        self.bar["value"] = value
        self.status.set(status)
        self.detail.set(detail)
        self.percent.set(f"{value:.0f}%")
        self.root.update_idletasks()
        self.root.update()

    def close(self) -> None:
        try:
            self.root.destroy()
        except tk.TclError:
            pass


def main() -> int:
    print(f"RS422 Excel Extractor {EXTRACTOR_VERSION}")
    selected = select_inputs()
    if selected is None:
        return 0
    logs_root = selected.resolve()
    output = logs_root / "EM-EQM_Thermal_Summary.xlsx"
    progress = ProgressDialog()
    try:
        progress.set(2, "Loading packet decoder", "Using the project EB packet utility.")
        decoder = import_decoder(None)
        progress.set(5, "Searching for RS422 logs", str(logs_root))
        logs = discover_logs(logs_root, None, None)
        if not logs:
            progress.close()
            messagebox.showwarning(
                "No RS422 logs found",
                "No files matching RS422if_*.log were found in that folder or its subfolders.",
            )
            return 1
        progress.set(10, "Loading chamber and ROV temperature data", str(logs_root))
        metadata = discover_temperature_metadata(logs_root)
        progress.set(
            20,
            "Preparing offline state checks",
            "State conditions will be detected from HK and matched to CH4/EB PSU current.",
        )
        print(
            "Metadata samples: "
            f"EB={len(metadata['eb_chamber_temp'])}, "
            f"OB={len(metadata['ob_chamber_temp'])}, "
            f"ROV={len(metadata['rov_therm'])}"
        )
        missing_sources = [
            label
            for label, key in (
                ("MSC2 / EB", "eb_chamber_temp"),
                ("MSC1 / OB", "ob_chamber_temp"),
                ("TP1000 / ROV", "rov_therm"),
            )
            if not metadata[key]
        ]
        if missing_sources:
            messagebox.showwarning(
                "Metadata folders not found",
                "The selected logs folder did not contain usable data for:\n"
                + "\n".join(f"• {label}" for label in missing_sources),
            )
        rows: list[tuple[Path, Summary, dict[str, Any]]] = []
        for index, log in enumerate(logs, 1):
            progress.set(
                25 + 63 * (index - 1) / max(1, len(logs)),
                f"Processing RS422 log {index} of {len(logs)}",
                str(log.relative_to(logs_root)),
            )
            result = summarise(log, decoder, "median")
            ch3_turn_on = find_channel_turn_on(logs_root, log, result.first_time, result.last_time, "CH3")
            ch4_turn_on = (
                None
                if ch3_turn_on is not None
                else find_channel_turn_on(logs_root, log, result.first_time, result.last_time, "CH4")
            )
            temperature_reference_time = ch3_turn_on or ch4_turn_on
            temperature_reference_channel = (
                "CH3" if ch3_turn_on is not None else "CH4" if ch4_turn_on is not None else None
            )
            if temperature_reference_time is None:
                result.warnings.append("CH3/CH4 turn-on transition not found; chamber temperatures unavailable")
            meta = metadata_for(log, result, metadata, temperature_reference_time)
            state_currents = offline_state_current_metadata(logs_root, log, result)
            meta.update(state_currents)
            # With no explicit transition timestamp, the first observed powered-OB
            # HK packet is the best reproducible proxy for switch-on.
            result.values["CH4_I at OB Switch on (A)"] = nearest_psu_current(
                logs_root, log, result.ob_switch_time or result.first_time
            )
            rows.append((log, result, meta))
            print(
                f"[{index}/{len(logs)}] {log.relative_to(logs_root)}: "
                f"HK={result.hk_count}, SCI={result.science_packets}, "
                f"EB={meta.get('EB Chamber Temp (°C)')}, "
                f"OB={meta.get('OB Chamber Temp(°C)')}, "
                f"ROV={meta.get('ROV_THERM (°C)')}, "
                f"States={len(result.state_condition_times)}/{len(STATE_CHECK_LABELS)}, "
                f"Temp anchor={temperature_reference_channel}@{temperature_reference_time}, "
                f"warnings={'; '.join(dict.fromkeys(result.warnings)) or 'none'}"
            )
        ungrouped_count = len(rows)
        progress.set(
            90,
            "Grouping SFT results",
            f"Selecting the preferred test for each EB/OB matrix point from {ungrouped_count} extracted rows.",
        )
        rows = group_rows_by_test_matrix(rows)
        print(f"Matrix grouping retained {len(rows)} row(s) from {ungrouped_count} extracted RS422 log row(s).")
        progress.set(
            96,
            "Writing Excel workbook",
            str(output.resolve()),
        )
        write_workbook(output.resolve(), rows)
        detected_state_count = sum(
            sum(1 for label in STATE_CHECK_LABELS if row_summary.values.get(f"{label} Current (mA)") is not None)
            for _row_log, row_summary, _row_metadata in rows
        )
        state_note = f"Recovered {detected_state_count} state-current reading(s) from HK and PSU data."
        progress.set(100, "Extraction complete", state_note)
    except Exception as exc:
        progress.close()
        messagebox.showerror("RS422 import failed", str(exc))
        return 1

    progress.close()
    messagebox.showinfo(
        "RS422 import complete",
        f"Created {len(rows)} summary row(s) in:\n{output.resolve()}",
    )
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
