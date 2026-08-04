#!/usr/bin/env python3
"""Interactive one/two-Y-axis plotter for an RS422 thermal-summary workbook.

Run without arguments and select the populated workbook.  Only the first
worksheet is read.  Any numeric column can be plotted on the left or optional
right Y axis against any populated X-axis column.
"""

from __future__ import annotations

import math
import sys
import tkinter as tk
import warnings
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from tkinter import filedialog, messagebox, ttk
from typing import Any, Iterable

from matplotlib.backends.backend_tkagg import FigureCanvasTkAgg, NavigationToolbar2Tk
from matplotlib.figure import Figure
from openpyxl import load_workbook


PLOTTER_VERSION = "2026.07.31-add-series-temperature-groups"


@dataclass
class WorkbookData:
    path: Path
    sheet_name: str
    headers: list[str]
    rows: list[dict[str, Any]]


@dataclass
class PlotRequest:
    x_header: str
    left_headers: list[str]
    right_headers: list[str]
    right_axis: bool = False
    plot_style: str = "Lines + markers"
    sort_x: bool = True
    annotate: bool = False
    label_header: str = "run"
    grid: bool = True
    left_scale: str = "linear"
    right_scale: str = "linear"
    title: str = ""
    sft_header: str | None = None
    included_sfts: frozenset[str] | None = None
    series_specs: tuple["SeriesSpec", ...] = ()


@dataclass(frozen=True)
class SeriesSpec:
    y_header: str
    side: str
    included_sfts: frozenset[str] | None
    legend_label: str
    group_header: str | None = None
    group_value: str | None = None


@dataclass
class PreparedXAxis:
    kind: str
    category_positions: dict[str, float]
    category_labels: list[str]


def finite_number(value: Any) -> float | None:
    if value is None or isinstance(value, bool):
        return None
    try:
        number = float(str(value).strip()) if isinstance(value, str) else float(value)
    except (TypeError, ValueError):
        return None
    return number if math.isfinite(number) else None


def nonblank(value: Any) -> bool:
    return value is not None and (not isinstance(value, str) or bool(value.strip()))


def display_value(value: Any) -> str:
    if isinstance(value, datetime):
        return value.strftime("%Y-%m-%d %H:%M:%S")
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, float):
        return f"{value:.8g}"
    return str(value)


def load_summary_workbook(path: Path) -> WorkbookData:
    """Read values from the first worksheet without altering the workbook."""
    warning_context = warnings.catch_warnings()
    warning_context.__enter__()
    warnings.simplefilter("ignore", UserWarning)
    workbook = None
    try:
        workbook = load_workbook(path, read_only=True, data_only=True)
        worksheet = workbook.worksheets[0]
        raw_headers = next(worksheet.iter_rows(min_row=1, max_row=1, values_only=True), ())
        headers: list[str] = []
        used_names: dict[str, int] = {}
        for index, value in enumerate(raw_headers, start=1):
            base = str(value).strip() if nonblank(value) else f"Column {index}"
            used_names[base] = used_names.get(base, 0) + 1
            headers.append(base if used_names[base] == 1 else f"{base} [{used_names[base]}]")

        rows: list[dict[str, Any]] = []
        for values in worksheet.iter_rows(min_row=2, max_col=len(headers), values_only=True):
            if not any(nonblank(value) for value in values):
                continue
            rows.append({header: value for header, value in zip(headers, values)})
        if not headers:
            raise ValueError("The first worksheet has no header row.")
        if not rows:
            raise ValueError("The first worksheet contains no populated data rows.")
        return WorkbookData(path.resolve(), worksheet.title, headers, rows)
    finally:
        if workbook is not None:
            workbook.close()
        warning_context.__exit__(None, None, None)


def numeric_headers(data: WorkbookData) -> list[str]:
    result: list[str] = []
    for header in data.headers:
        populated = [row.get(header) for row in data.rows if nonblank(row.get(header))]
        if populated and any(finite_number(value) is not None for value in populated):
            result.append(header)
    return result


def populated_headers(data: WorkbookData) -> list[str]:
    return [header for header in data.headers if any(nonblank(row.get(header)) for row in data.rows)]


def row_is_included(
    row: dict[str, Any],
    sft_header: str | None,
    included_sfts: frozenset[str] | None,
    group_header: str | None = None,
    group_value: str | None = None,
) -> bool:
    if sft_header is None or included_sfts is None:
        sft_matches = True
    else:
        sft_matches = display_value(row.get(sft_header)) in included_sfts
    if not sft_matches:
        return False
    if group_header is None or group_value is None:
        return True
    return display_value(row.get(group_header)) == group_value


def prepare_x_axis(
    data: WorkbookData,
    header: str,
    *,
    sft_header: str | None = None,
    included_sfts: frozenset[str] | None = None,
) -> PreparedXAxis:
    values = [
        row.get(header)
        for row in data.rows
        if row_is_included(row, sft_header, included_sfts) and nonblank(row.get(header))
    ]
    if not values:
        raise ValueError(f"{header!r} has no populated X-axis values.")
    if all(isinstance(value, (datetime, date)) for value in values):
        return PreparedXAxis("datetime", {}, [])
    if all(finite_number(value) is not None for value in values):
        return PreparedXAxis("numeric", {}, [])

    labels: list[str] = []
    positions: dict[str, float] = {}
    for value in values:
        label = display_value(value)
        if label not in positions:
            positions[label] = float(len(labels))
            labels.append(label)
    return PreparedXAxis("categorical", positions, labels)


def converted_x(value: Any, prepared: PreparedXAxis) -> Any | None:
    if not nonblank(value):
        return None
    if prepared.kind == "numeric":
        return finite_number(value)
    if prepared.kind == "datetime":
        return value if isinstance(value, (datetime, date)) else None
    return prepared.category_positions.get(display_value(value))


def build_plot_points(
    data: WorkbookData,
    x_header: str,
    y_header: str,
    prepared_x: PreparedXAxis,
    *,
    label_header: str,
    sort_x: bool,
    log_scale: bool,
    sft_header: str | None = None,
    included_sfts: frozenset[str] | None = None,
    group_header: str | None = None,
    group_value: str | None = None,
) -> list[tuple[Any, float, str, int]]:
    points: list[tuple[Any, float, str, int]] = []
    for row_number, row in enumerate(data.rows, start=2):
        if not row_is_included(
            row,
            sft_header,
            included_sfts,
            group_header,
            group_value,
        ):
            continue
        x_value = converted_x(row.get(x_header), prepared_x)
        y_value = finite_number(row.get(y_header))
        if x_value is None or y_value is None or (log_scale and y_value <= 0):
            continue
        label = display_value(row.get(label_header)) if nonblank(row.get(label_header)) else f"Row {row_number}"
        points.append((x_value, y_value, label, row_number))
    if sort_x and prepared_x.kind in ("numeric", "datetime"):
        points.sort(key=lambda point: point[0])
    return points


def _style_args(style: str, series_index: int) -> dict[str, Any]:
    markers = ("o", "s", "^", "D", "v", "P", "X", "<", ">", "h")
    common: dict[str, Any] = {
        "linewidth": 1.5,
        "markersize": 5.5,
        "marker": markers[series_index % len(markers)],
        "picker": 5,
    }
    if style == "Markers only":
        common["linestyle"] = "None"
    elif style == "Lines only":
        common["marker"] = None
    else:
        common["linestyle"] = "-"
    return common


def render_plot(
    figure: Figure,
    data: WorkbookData,
    request: PlotRequest,
) -> tuple[list[Any], list[str]]:
    """Render a request and return pickable artists plus status messages."""
    if request.series_specs:
        series_specs = list(request.series_specs)
    else:
        series_specs = [SeriesSpec(header, "left", request.included_sfts, header) for header in request.left_headers]
        if request.right_axis:
            series_specs.extend(
                SeriesSpec(header, "right", request.included_sfts, header) for header in request.right_headers
            )

    if not series_specs:
        raise ValueError("Select at least one Y-axis parameter.")
    if request.sft_header is not None and any(
        spec.included_sfts is not None and not spec.included_sfts for spec in series_specs
    ):
        raise ValueError("Select at least one SFT to include.")

    filtered_sets = [spec.included_sfts for spec in series_specs if spec.included_sfts is not None]
    x_included_sfts = None if len(filtered_sets) != len(series_specs) else frozenset().union(*filtered_sets)

    prepared_x = prepare_x_axis(
        data,
        request.x_header,
        sft_header=request.sft_header,
        included_sfts=x_included_sfts,
    )
    figure.clear()
    left_axis = figure.add_subplot(111)
    right_axis = left_axis.twinx() if any(spec.side == "right" for spec in series_specs) else None
    colors = (
        "#0072B2",
        "#D55E00",
        "#009E73",
        "#CC79A7",
        "#E69F00",
        "#56B4E9",
        "#F0E442",
        "#000000",
        "#6A3D9A",
        "#A6761D",
    )
    artists: list[Any] = []
    messages: list[str] = []

    def draw(axis: Any, specs: Iterable[SeriesSpec], scale: str, offset: int, side: str) -> None:
        for series_index, spec in enumerate(specs, start=offset):
            points = build_plot_points(
                data,
                request.x_header,
                spec.y_header,
                prepared_x,
                label_header=request.label_header,
                sort_x=request.sort_x,
                log_scale=scale == "log",
                sft_header=request.sft_header,
                included_sfts=spec.included_sfts,
                group_header=spec.group_header,
                group_value=spec.group_value,
            )
            if not points:
                messages.append(f"{side}: {spec.legend_label} — no usable paired values")
                continue
            xs = [point[0] for point in points]
            ys = [point[1] for point in points]
            (artist,) = axis.plot(
                xs,
                ys,
                label=spec.legend_label,
                color=colors[series_index % len(colors)],
                **_style_args(request.plot_style, series_index),
            )
            artist._summary_plotter_points = points  # type: ignore[attr-defined]
            artist._summary_plotter_header = spec.legend_label  # type: ignore[attr-defined]
            artist._summary_plotter_x_header = request.x_header  # type: ignore[attr-defined]
            artists.append(artist)
            messages.append(f"{side}: {spec.legend_label} — {len(points)} point(s)")
            if request.annotate:
                for x_value, y_value, label, _row_number in points:
                    axis.annotate(
                        label,
                        (x_value, y_value),
                        xytext=(4, 4),
                        textcoords="offset points",
                        fontsize=7,
                        alpha=0.8,
                    )
        axis.set_yscale(scale)

    left_specs = [spec for spec in series_specs if spec.side == "left"]
    right_specs = [spec for spec in series_specs if spec.side == "right"]
    draw(left_axis, left_specs, request.left_scale, 0, "Left")
    if right_axis is not None:
        draw(right_axis, right_specs, request.right_scale, len(left_specs), "Right")

    left_axis.set_xlabel(request.x_header)
    left_headers = list(dict.fromkeys(spec.y_header for spec in left_specs))
    right_headers = list(dict.fromkeys(spec.y_header for spec in right_specs))
    if len(left_headers) == 1:
        left_axis.set_ylabel(left_headers[0], color="#0072B2")
    elif left_headers:
        left_axis.set_ylabel("Left-axis parameters", color="#0072B2")
    if right_axis is not None:
        if len(right_headers) == 1:
            right_axis.set_ylabel(right_headers[0], color="#D55E00")
        else:
            right_axis.set_ylabel("Right-axis parameters", color="#D55E00")

    if prepared_x.kind == "categorical":
        positions = list(range(len(prepared_x.category_labels)))
        left_axis.set_xticks(positions)
        left_axis.set_xticklabels(prepared_x.category_labels, rotation=45, ha="right")
    elif prepared_x.kind == "datetime":
        figure.autofmt_xdate(rotation=35)

    left_axis.grid(request.grid, color="#D9D9D9", linewidth=0.7, alpha=0.75)
    title = request.title.strip() or f"Selected parameters vs {request.x_header}"
    left_axis.set_title(title)

    handles = list(left_axis.get_lines())
    if right_axis is not None:
        handles.extend(right_axis.get_lines())
    if handles:
        left_axis.legend(
            handles=handles,
            labels=[handle.get_label() for handle in handles],
            loc="best",
            fontsize=8,
            framealpha=0.9,
        )
    figure.tight_layout()
    return artists, messages


def is_temperature_header(header: str) -> bool:
    text = header.lower()
    return "temp" in text or "therm" in text


def is_voltage_or_current_header(header: str) -> bool:
    text = header.lower()
    return "(v)" in text or "voltage" in text or "rail" in text or "(a)" in text or "current" in text


class ParameterList(ttk.LabelFrame):
    def __init__(self, master: tk.Misc, title: str) -> None:
        super().__init__(master, text=title)
        self.options: list[str] = []
        self.visible: list[str] = []
        self.selected: set[str] = set()
        self.search_var = tk.StringVar()

        search_row = ttk.Frame(self)
        search_row.pack(fill="x", padx=6, pady=(5, 3))
        ttk.Label(search_row, text="Filter:").pack(side="left")
        ttk.Entry(search_row, textvariable=self.search_var).pack(side="left", fill="x", expand=True, padx=(5, 0))

        list_frame = ttk.Frame(self)
        list_frame.pack(fill="both", expand=True, padx=6)
        self.listbox = tk.Listbox(
            list_frame,
            selectmode="extended",
            exportselection=False,
            height=12,
        )
        scrollbar = ttk.Scrollbar(list_frame, orient="vertical", command=self.listbox.yview)
        self.listbox.configure(yscrollcommand=scrollbar.set)
        self.listbox.pack(side="left", fill="both", expand=True)
        scrollbar.pack(side="right", fill="y")

        button_row = ttk.Frame(self)
        button_row.pack(fill="x", padx=6, pady=5)
        ttk.Button(button_row, text="All shown", command=self.select_all_visible).pack(side="left")
        ttk.Button(button_row, text="Clear shown", command=self.clear_visible).pack(side="left", padx=5)

        self.search_var.trace_add("write", self._search_changed)
        self.listbox.bind("<<ListboxSelect>>", self._selection_changed)

    def set_options(self, options: Iterable[str]) -> None:
        self._sync_visible_selection()
        self.options = list(options)
        self.selected.intersection_update(self.options)
        self._refresh()

    def set_selected(self, options: Iterable[str]) -> None:
        self.selected = set(options).intersection(self.options)
        self._refresh()

    def get_selected(self) -> list[str]:
        self._sync_visible_selection()
        return [option for option in self.options if option in self.selected]

    def _sync_visible_selection(self) -> None:
        selected_indices = set(self.listbox.curselection())
        for index, option in enumerate(self.visible):
            if index in selected_indices:
                self.selected.add(option)
            else:
                self.selected.discard(option)

    def _selection_changed(self, _event: Any = None) -> None:
        self._sync_visible_selection()

    def _search_changed(self, *_args: Any) -> None:
        self._sync_visible_selection()
        self._refresh()

    def _refresh(self) -> None:
        query = self.search_var.get().strip().lower()
        self.visible = [option for option in self.options if query in option.lower()]
        self.listbox.delete(0, "end")
        for index, option in enumerate(self.visible):
            self.listbox.insert("end", option)
            if option in self.selected:
                self.listbox.selection_set(index)

    def select_all_visible(self) -> None:
        self.selected.update(self.visible)
        self._refresh()

    def clear_visible(self) -> None:
        self.selected.difference_update(self.visible)
        self._refresh()


class ThermalSummaryPlotter:
    def __init__(self, root: tk.Tk) -> None:
        self.root = root
        self.root.title(f"Thermal Summary Plotter — {PLOTTER_VERSION}")
        self.root.geometry("1320x860")
        self.root.minsize(1050, 680)
        self.data: WorkbookData | None = None
        self.pick_annotation: Any | None = None
        self.sft_header: str | None = None
        self.sft_options: list[str] = []
        self.included_sfts: set[str] = set()
        self.plotted_series: list[SeriesSpec] = []
        self.plotted_x_header: str | None = None

        self.x_var = tk.StringVar()
        self.label_var = tk.StringVar(value="run")
        self.right_enabled = tk.BooleanVar(value=True)
        self.style_var = tk.StringVar(value="Lines + markers")
        self.sort_var = tk.BooleanVar(value=True)
        self.annotate_var = tk.BooleanVar(value=False)
        self.grid_var = tk.BooleanVar(value=True)
        self.left_scale_var = tk.StringVar(value="linear")
        self.right_scale_var = tk.StringVar(value="linear")
        self.group_var = tk.StringVar(value="None")
        self.title_var = tk.StringVar()
        self.series_label_var = tk.StringVar()
        self.file_var = tk.StringVar(value="No workbook loaded")
        self.status_var = tk.StringVar(value="Open a populated thermal-summary workbook.")
        self.sft_count_var = tk.StringVar(value="SFTs: no workbook loaded")
        self.series_count_var = tk.StringVar(value="Plotted series: 0")

        self._build_ui()

    def _build_ui(self) -> None:
        file_bar = ttk.Frame(self.root, padding=(8, 7))
        file_bar.pack(fill="x")
        ttk.Button(file_bar, text="Open workbook…", command=self.open_workbook).pack(side="left")
        ttk.Label(file_bar, textvariable=self.file_var).pack(side="left", fill="x", expand=True, padx=10)
        ttk.Label(file_bar, text=PLOTTER_VERSION).pack(side="right")

        main = ttk.Panedwindow(self.root, orient="horizontal")
        main.pack(fill="both", expand=True, padx=8, pady=(0, 8))
        controls = ttk.Frame(main, padding=7)
        plot_area = ttk.Frame(main)
        main.add(controls, weight=1)
        main.add(plot_area, weight=3)

        axis_frame = ttk.LabelFrame(controls, text="Axes", padding=7)
        axis_frame.pack(fill="x")
        ttk.Label(axis_frame, text="X-axis parameter").grid(row=0, column=0, sticky="w")
        self.x_combo = ttk.Combobox(axis_frame, textvariable=self.x_var, state="readonly")
        self.x_combo.grid(row=1, column=0, columnspan=2, sticky="ew", pady=(2, 6))
        ttk.Label(axis_frame, text="Point-label parameter").grid(row=2, column=0, sticky="w")
        self.label_combo = ttk.Combobox(axis_frame, textvariable=self.label_var, state="readonly")
        self.label_combo.grid(row=3, column=0, columnspan=2, sticky="ew", pady=(2, 0))
        axis_frame.columnconfigure(0, weight=1)

        preset_frame = ttk.LabelFrame(controls, text="Quick selection", padding=7)
        preset_frame.pack(fill="x", pady=(7, 0))
        ttk.Button(
            preset_frame,
            text="Temperatures left; volts/current right",
            command=self.apply_temperature_voltage_preset,
        ).pack(fill="x")
        ttk.Button(preset_frame, text="Clear both Y axes", command=self.clear_y_axes).pack(fill="x", pady=(5, 0))
        self.sft_button = ttk.Button(
            preset_frame,
            text="Select SFTs…",
            command=self.select_sfts,
            state="disabled",
        )
        self.sft_button.pack(fill="x", pady=(5, 0))
        ttk.Label(preset_frame, textvariable=self.sft_count_var).pack(anchor="w", pady=(4, 0))

        selectors = ttk.Panedwindow(controls, orient="vertical")
        selectors.pack(fill="both", expand=True, pady=(7, 0))
        self.left_list = ParameterList(selectors, "Left Y axis — select one or more")
        self.right_list = ParameterList(selectors, "Right Y axis — select one or more")
        selectors.add(self.left_list, weight=1)
        selectors.add(self.right_list, weight=1)

        options = ttk.LabelFrame(controls, text="Plot options", padding=7)
        options.pack(fill="x", pady=(7, 0))
        ttk.Checkbutton(options, text="Enable right Y axis", variable=self.right_enabled).grid(
            row=0, column=0, columnspan=2, sticky="w"
        )
        ttk.Checkbutton(options, text="Sort numeric/date X", variable=self.sort_var).grid(row=1, column=0, sticky="w")
        ttk.Checkbutton(options, text="Grid", variable=self.grid_var).grid(row=1, column=1, sticky="w")
        ttk.Checkbutton(options, text="Annotate point labels", variable=self.annotate_var).grid(
            row=2, column=0, columnspan=2, sticky="w"
        )
        ttk.Label(options, text="Style").grid(row=3, column=0, sticky="w", pady=(5, 0))
        ttk.Combobox(
            options,
            textvariable=self.style_var,
            values=("Lines + markers", "Markers only", "Lines only"),
            state="readonly",
            width=17,
        ).grid(row=3, column=1, sticky="ew", pady=(5, 0))
        ttk.Label(options, text="Left scale").grid(row=4, column=0, sticky="w")
        ttk.Combobox(
            options, textvariable=self.left_scale_var, values=("linear", "log"), state="readonly", width=9
        ).grid(row=4, column=1, sticky="ew")
        ttk.Label(options, text="Right scale").grid(row=5, column=0, sticky="w")
        ttk.Combobox(
            options, textvariable=self.right_scale_var, values=("linear", "log"), state="readonly", width=9
        ).grid(row=5, column=1, sticky="ew")
        ttk.Label(options, text="Group series by").grid(row=6, column=0, sticky="w", pady=(5, 0))
        ttk.Combobox(
            options,
            textvariable=self.group_var,
            values=("None", "EB temperature", "OB temperature"),
            state="readonly",
            width=17,
        ).grid(row=6, column=1, sticky="ew", pady=(5, 0))
        ttk.Label(options, text="Title (optional)").grid(row=7, column=0, columnspan=2, sticky="w", pady=(5, 0))
        ttk.Entry(options, textvariable=self.title_var).grid(row=8, column=0, columnspan=2, sticky="ew")
        ttk.Label(options, text="New-series label (optional)").grid(
            row=9, column=0, columnspan=2, sticky="w", pady=(5, 0)
        )
        ttk.Entry(options, textvariable=self.series_label_var).grid(row=10, column=0, columnspan=2, sticky="ew")
        options.columnconfigure(1, weight=1)

        action_row = ttk.Frame(controls)
        action_row.pack(fill="x", pady=(7, 0))
        ttk.Button(action_row, text="Plot new", command=self.plot).pack(side="left", fill="x", expand=True)
        ttk.Button(action_row, text="Add series", command=self.add_series).pack(
            side="left", fill="x", expand=True, padx=(6, 0)
        )
        second_action_row = ttk.Frame(controls)
        second_action_row.pack(fill="x", pady=(5, 0))
        ttk.Button(second_action_row, text="Undo last", command=self.undo_last_series).pack(
            side="left", fill="x", expand=True
        )
        ttk.Button(second_action_row, text="Save plot…", command=self.save_plot).pack(
            side="left", fill="x", expand=True, padx=(6, 0)
        )
        ttk.Label(controls, textvariable=self.series_count_var).pack(anchor="w", pady=(4, 0))

        self.figure = Figure(figsize=(9.3, 7.0), dpi=100)
        self.canvas = FigureCanvasTkAgg(self.figure, master=plot_area)
        toolbar = NavigationToolbar2Tk(self.canvas, plot_area, pack_toolbar=False)
        toolbar.update()
        toolbar.pack(side="top", fill="x")
        self.canvas.get_tk_widget().pack(fill="both", expand=True)
        self.canvas.mpl_connect("pick_event", self.on_pick)

        status = ttk.Label(self.root, textvariable=self.status_var, anchor="w", relief="sunken", padding=(7, 4))
        status.pack(fill="x", side="bottom")

    def open_workbook(self, path: str | Path | None = None) -> None:
        if path is None:
            chosen = filedialog.askopenfilename(
                parent=self.root,
                title="Select populated thermal-summary workbook",
                filetypes=(("Excel workbooks", "*.xlsx"), ("All files", "*.*")),
            )
            if not chosen:
                return
            path = chosen
        try:
            data = load_summary_workbook(Path(path))
        except Exception as exc:
            messagebox.showerror("Workbook could not be loaded", str(exc), parent=self.root)
            return

        self.data = data
        self.plotted_series = []
        self.plotted_x_header = None
        self._update_series_count()
        x_headers = populated_headers(data)
        y_headers = numeric_headers(data)
        self.x_combo["values"] = x_headers
        self.label_combo["values"] = x_headers
        self.left_list.set_options(y_headers)
        self.right_list.set_options(y_headers)
        self.sft_header = next(
            (header for header in data.headers if header.strip().lower() == "run"),
            None,
        )
        if self.sft_header is not None:
            self.sft_options = list(
                dict.fromkeys(
                    display_value(row.get(self.sft_header)) for row in data.rows if nonblank(row.get(self.sft_header))
                )
            )
        else:
            self.sft_options = []
        self.included_sfts = set(self.sft_options)
        if self.sft_options:
            self.sft_button.state(["!disabled"])
        else:
            self.sft_button.state(["disabled"])
        self._update_sft_count()
        preferred_x = next(
            (
                header
                for header in ("Matrix EB Setpoint (°C)", "Temp (°C)", "EB Chamber Temp (°C)")
                if header in x_headers
            ),
            x_headers[0],
        )
        self.x_var.set(preferred_x)
        self.label_var.set("run" if "run" in x_headers else x_headers[0])
        self.file_var.set(f"{data.path.name} — first tab: {data.sheet_name}")
        self.status_var.set(f"Loaded {len(data.rows)} row(s); {len(y_headers)} numeric Y parameter(s).")
        self.apply_temperature_voltage_preset()
        self.plot()

    def apply_temperature_voltage_preset(self) -> None:
        if self.data is None:
            return
        numeric = numeric_headers(self.data)
        temperatures = [header for header in numeric if is_temperature_header(header) and header != self.x_var.get()]
        electrical = [
            header for header in numeric if is_voltage_or_current_header(header) and header != self.x_var.get()
        ]
        self.left_list.set_selected(temperatures)
        self.right_list.set_selected(electrical)
        self.right_enabled.set(bool(electrical))

    def clear_y_axes(self) -> None:
        self.left_list.set_selected([])
        self.right_list.set_selected([])

    def _update_sft_count(self) -> None:
        if not self.sft_options:
            self.sft_count_var.set("SFTs: no run column found")
            return
        self.sft_count_var.set(f"SFTs included: {len(self.included_sfts)} / {len(self.sft_options)}")

    def select_sfts(self) -> None:
        if not self.sft_options:
            messagebox.showinfo(
                "Select SFTs",
                "The first worksheet does not contain a populated run column.",
                parent=self.root,
            )
            return

        window = tk.Toplevel(self.root)
        window.title("Select SFTs to include")
        window.geometry("520x610")
        window.transient(self.root)

        ttk.Label(
            window,
            text="Select the SFT/run rows used by every plotted series:",
            padding=(10, 10, 10, 4),
        ).pack(anchor="w")
        selector = ParameterList(window, "Available SFTs")
        selector.pack(fill="both", expand=True, padx=10, pady=(0, 8))
        selector.set_options(self.sft_options)
        selector.set_selected(self.included_sfts)

        def apply_selection() -> None:
            selected = selector.get_selected()
            if not selected:
                messagebox.showwarning(
                    "Select SFTs",
                    "Select at least one SFT.",
                    parent=window,
                )
                return
            self.included_sfts = set(selected)
            self._update_sft_count()
            window.destroy()
            self.status_var.set(
                "SFT selection updated. Click Add series to overlay it, or Plot new to replace the existing plot."
            )

        button_row = ttk.Frame(window, padding=(10, 0, 10, 10))
        button_row.pack(fill="x")
        ttk.Button(button_row, text="Apply", command=apply_selection).pack(side="left")
        ttk.Button(button_row, text="Cancel", command=window.destroy).pack(side="left", padx=(7, 0))
        window.bind("<Return>", lambda _event: apply_selection())
        window.bind("<Escape>", lambda _event: window.destroy())
        window.grab_set()
        window.focus_set()

    def current_request(self) -> PlotRequest:
        return PlotRequest(
            x_header=self.x_var.get(),
            left_headers=self.left_list.get_selected(),
            right_headers=self.right_list.get_selected(),
            right_axis=self.right_enabled.get(),
            plot_style=self.style_var.get(),
            sort_x=self.sort_var.get(),
            annotate=self.annotate_var.get(),
            label_header=self.label_var.get(),
            grid=self.grid_var.get(),
            left_scale=self.left_scale_var.get(),
            right_scale=self.right_scale_var.get(),
            title=self.title_var.get(),
            sft_header=self.sft_header,
            included_sfts=(frozenset(self.included_sfts) if self.sft_header is not None else None),
        )

    def _sft_suffix(self, selected: frozenset[str] | None) -> str:
        if selected is None or set(selected) == set(self.sft_options):
            return ""
        ordered = [name for name in self.sft_options if name in selected]
        if len(ordered) <= 3:
            return ", ".join(ordered)
        return f"{len(ordered)} selected SFTs"

    def _temperature_group_header(self) -> tuple[str | None, str]:
        if self.data is None or self.group_var.get() == "None":
            return None, ""
        if self.group_var.get() == "EB temperature":
            candidates = (
                "Matrix EB Setpoint (°C)",
                "EB Chamber Temp (°C)",
                "Temp (°C)",
            )
            prefix = "EB"
        else:
            candidates = (
                "Matrix OB Setpoint (°C)",
                "OB Chamber Temp(°C)",
            )
            prefix = "OB"
        header = next((name for name in candidates if name in self.data.headers), None)
        if header is None:
            raise ValueError(f"No usable {prefix} temperature/setpoint column was found.")
        return header, prefix

    @staticmethod
    def _temperature_group_label(value: str) -> str:
        number = finite_number(value)
        return f"{number:g} °C" if number is not None else value

    def _series_from_controls(self) -> list[SeriesSpec]:
        request = self.current_request()
        selected_sfts = request.included_sfts
        custom_label = self.series_label_var.get().strip()
        entries = [(header, "left") for header in request.left_headers]
        if request.right_axis:
            entries.extend((header, "right") for header in request.right_headers)
        if not entries:
            raise ValueError("Select at least one Y-axis parameter.")
        if self.sft_header is not None and selected_sfts is not None and not selected_sfts:
            raise ValueError("Select at least one SFT to include.")

        suffix = self._sft_suffix(selected_sfts)
        group_header, group_prefix = self._temperature_group_header()
        if group_header is None:
            group_values: list[str | None] = [None]
        else:
            group_values = list(
                dict.fromkeys(
                    display_value(row.get(group_header))
                    for row in self.data.rows
                    if row_is_included(row, self.sft_header, selected_sfts) and nonblank(row.get(group_header))
                )
            )
            if not group_values:
                raise ValueError(f"The selected SFTs have no {group_prefix} temperature values.")
        multiple = len(entries) * len(group_values) > 1
        specs: list[SeriesSpec] = []
        for header, side in entries:
            for group_value in group_values:
                if custom_label:
                    legend_label = f"{custom_label}: {header}" if multiple else custom_label
                else:
                    legend_label = header
                if group_value is not None:
                    legend_label += f" — {group_prefix} {self._temperature_group_label(group_value)}"
                if suffix:
                    legend_label += f" — {suffix}"
                specs.append(
                    SeriesSpec(
                        header,
                        side,
                        selected_sfts,
                        legend_label,
                        group_header,
                        group_value,
                    )
                )
        return specs

    def _update_series_count(self) -> None:
        self.series_count_var.set(f"Plotted series: {len(self.plotted_series)}")

    def _render_plotted_series(self) -> None:
        if self.data is None or not self.plotted_series or self.plotted_x_header is None:
            self.figure.clear()
            self.canvas.draw_idle()
            self._update_series_count()
            return
        request = self.current_request()
        request.x_header = self.plotted_x_header
        request.series_specs = tuple(self.plotted_series)
        _artists, messages = render_plot(self.figure, self.data, request)
        self.pick_annotation = None
        self.canvas.draw_idle()
        self._update_series_count()
        self.status_var.set(" | ".join(messages) if messages else "No usable series were plotted.")

    def plot(self) -> None:
        if self.data is None:
            return
        try:
            self.plotted_series = self._series_from_controls()
            self.plotted_x_header = self.x_var.get()
            self._render_plotted_series()
        except Exception as exc:
            messagebox.showwarning("Plot could not be created", str(exc), parent=self.root)
            return

    def add_series(self) -> None:
        if self.data is None:
            return
        if not self.plotted_series:
            self.plot()
            return
        if self.x_var.get() != self.plotted_x_header:
            messagebox.showwarning(
                "X axis changed",
                "Add series keeps the existing X axis. Restore the previous X axis, "
                "or click Plot new to start a plot with the new X axis.",
                parent=self.root,
            )
            return
        try:
            candidates = self._series_from_controls()
            existing = {
                (
                    spec.y_header,
                    spec.side,
                    spec.included_sfts,
                    spec.group_header,
                    spec.group_value,
                )
                for spec in self.plotted_series
            }
            additions = [
                spec
                for spec in candidates
                if (
                    spec.y_header,
                    spec.side,
                    spec.included_sfts,
                    spec.group_header,
                    spec.group_value,
                )
                not in existing
            ]
            if not additions:
                raise ValueError("Those parameters and SFTs are already present in the plot.")
            self.plotted_series.extend(additions)
            self._render_plotted_series()
        except Exception as exc:
            messagebox.showwarning("Series could not be added", str(exc), parent=self.root)

    def undo_last_series(self) -> None:
        if not self.plotted_series:
            return
        self.plotted_series.pop()
        if not self.plotted_series:
            self.plotted_x_header = None
        try:
            self._render_plotted_series()
        except Exception as exc:
            messagebox.showwarning("Plot could not be updated", str(exc), parent=self.root)

    def on_pick(self, event: Any) -> None:
        artist = event.artist
        points = getattr(artist, "_summary_plotter_points", None)
        indices = getattr(event, "ind", None)
        if not points or indices is None or len(indices) == 0:
            return
        index = int(indices[0])
        if index >= len(points):
            return
        x_value, y_value, label, row_number = points[index]
        axis = artist.axes
        if self.pick_annotation is not None:
            try:
                self.pick_annotation.remove()
            except Exception:
                pass
        series = getattr(artist, "_summary_plotter_header", "Series")
        x_header = getattr(artist, "_summary_plotter_x_header", "X")
        text = (
            f"{series}\n{x_header}: {display_value(x_value)}\n"
            f"Value: {display_value(y_value)}\n{label} — Excel row {row_number}"
        )
        self.pick_annotation = axis.annotate(
            text,
            (x_value, y_value),
            xytext=(12, 12),
            textcoords="offset points",
            bbox={"boxstyle": "round,pad=0.4", "facecolor": "white", "alpha": 0.95},
            arrowprops={"arrowstyle": "->", "color": "#555555"},
            fontsize=8,
            zorder=20,
        )
        self.canvas.draw_idle()

    def save_plot(self) -> None:
        if not self.figure.axes:
            messagebox.showinfo("Save plot", "Create a plot first.", parent=self.root)
            return
        chosen = filedialog.asksaveasfilename(
            parent=self.root,
            title="Save plot",
            defaultextension=".png",
            initialfile="thermal_summary_plot.png",
            filetypes=(
                ("PNG image", "*.png"),
                ("PDF document", "*.pdf"),
                ("SVG vector image", "*.svg"),
            ),
        )
        if not chosen:
            return
        try:
            self.figure.savefig(chosen, dpi=200, bbox_inches="tight")
        except Exception as exc:
            messagebox.showerror("Plot could not be saved", str(exc), parent=self.root)
            return
        self.status_var.set(f"Saved plot: {chosen}")


def main() -> int:
    print(f"Thermal Summary Plotter {PLOTTER_VERSION}")
    root = tk.Tk()
    app = ThermalSummaryPlotter(root)
    if len(sys.argv) > 1:
        root.after(100, lambda: app.open_workbook(sys.argv[1]))
    else:
        root.after(100, app.open_workbook)
    root.mainloop()
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
