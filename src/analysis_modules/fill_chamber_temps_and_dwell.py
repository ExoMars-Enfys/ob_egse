from __future__ import annotations

import csv
import math
import re
import sys
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path

from openpyxl import load_workbook

SRC_DIR = Path(__file__).resolve().parents[1]
if str(SRC_DIR) not in sys.path:
    sys.path.insert(0, str(SRC_DIR))

from scripts.thermal_summary import _find_ob_switch_on_index, _find_preferred_rs422_logs, _first_valid_sample_after_index
from scripts.analysis import _read_all_packets

THERMAL_ROOT = Path(r"C:\Users\GK\OneDrive - University College London\General - Enfys - Shared\Test\EM-EQM Thermal Test")
WORKBOOK_PATH = THERMAL_ROOT / "EM_Thermal Readings Summary.xlsx"
EB_CSV_PATH = THERMAL_ROOT / "EB Temp logs" / "20260217_165944_EB.csv"
OB_CSV_PATH = THERMAL_ROOT / "OB Temp logs" / "20260212_142809_OB.csv"
ROV_CSV_PATH = THERMAL_ROOT / "OB Temp logs" / "2026_02_18-104107_RovTherm.csv"

EB_DWELL_HEADER = "EB Dwell Time (s)"
OB_DWELL_HEADER = "OB Dwell Time (s)"

STABILITY_TOLERANCE_C = 1.0
MAX_NEAREST_DELTA_SECONDS = 6 * 3600
MAX_PLATEAU_SLOPE_C_PER_MIN = 0.08
REQUIRE_IN_BAND_NEAR_SWITCH_SECONDS = 60 * 60


@dataclass
class SeriesPoint:
    t: datetime
    value: float


def _parse_time(value: str) -> datetime | None:
    text = (value or "").strip()
    if not text:
        return None
    for fmt in ("%Y%m%d_%H%M%S", "%Y-%m-%d %H:%M:%S"):
        try:
            return datetime.strptime(text, fmt)
        except ValueError:
            continue
    return None


def _safe_float(value: str) -> float | None:
    text = (value or "").strip()
    if not text:
        return None
    try:
        val = float(text)
    except ValueError:
        return None
    if math.isnan(val) or math.isinf(val):
        return None
    return val


def _extract_sft_key(text: str) -> str | None:
    match = re.search(r"SFT\s*(\d+)", text or "", flags=re.IGNORECASE)
    if not match:
        return None
    return f"SFT{int(match.group(1))}"


def _parse_setpoint(temp_label: str) -> float | None:
    match = re.search(r"(-?\d+)\s*C", str(temp_label), flags=re.IGNORECASE)
    if not match:
        return None
    return float(match.group(1))


def _read_eb_series(path: Path) -> list[SeriesPoint]:
    points: list[SeriesPoint] = []
    with path.open("r", encoding="utf-8-sig", errors="ignore") as handle:
        reader = csv.reader(handle)
        header = next(reader, [])
        if not header:
            return points
        channel_index = None
        for idx, name in enumerate(header):
            if (name or "").strip().upper() == "CH6":
                channel_index = idx
                break
        if channel_index is None:
            return points

        for row in reader:
            if not row or len(row) <= channel_index:
                continue
            t = _parse_time(row[0])
            value = _safe_float(row[channel_index])
            if t is None or value is None:
                continue
            points.append(SeriesPoint(t=t, value=value))

    return points


def _read_ob_series(path: Path) -> list[SeriesPoint]:
    points: list[SeriesPoint] = []
    with path.open("r", encoding="utf-8-sig", errors="ignore") as handle:
        reader = csv.reader(handle)
        for row in reader:
            if not row or len(row) < 2:
                continue
            t = _parse_time(row[0])
            value = _safe_float(row[1])
            if t is None or value is None:
                continue
            points.append(SeriesPoint(t=t, value=value))
    return points


def _pt1000_resistance_from_temp(temp_c: float) -> float:
    r0 = 1000.0
    a = 3.9083e-3
    b = -5.775e-7
    c = -4.183e-12
    if temp_c >= 0:
        return r0 * (1 + a * temp_c + b * temp_c * temp_c)
    return r0 * (1 + a * temp_c + b * temp_c * temp_c + c * (temp_c - 100) * temp_c**3)


def _build_pt1000_lookup(min_c: float = -80.0, max_c: float = 80.0, step_c: float = 0.1) -> list[tuple[float, float]]:
    table: list[tuple[float, float]] = []
    steps = int(round((max_c - min_c) / step_c))
    for idx in range(steps + 1):
        temp_c = min_c + idx * step_c
        table.append((_pt1000_resistance_from_temp(temp_c), temp_c))
    return sorted(table, key=lambda x: x[0])


def _pt1000_lookup_temp_from_ohms(resistance: float, lookup_table: list[tuple[float, float]]) -> float | None:
    if not lookup_table:
        return None
    if resistance <= lookup_table[0][0]:
        return lookup_table[0][1]
    if resistance >= lookup_table[-1][0]:
        return lookup_table[-1][1]

    low = 0
    high = len(lookup_table) - 1
    while high - low > 1:
        mid = (low + high) // 2
        if lookup_table[mid][0] <= resistance:
            low = mid
        else:
            high = mid

    r1, t1 = lookup_table[low]
    r2, t2 = lookup_table[high]
    if r2 == r1:
        return t1
    fraction = (resistance - r1) / (r2 - r1)
    return t1 + fraction * (t2 - t1)


def _read_rov_series(path: Path) -> list[SeriesPoint]:
    points: list[SeriesPoint] = []
    lookup = _build_pt1000_lookup()
    with path.open("r", encoding="utf-8-sig", errors="ignore") as handle:
        reader = csv.reader(handle)
        header = next(reader, [])
        if not header:
            return points

        time_idx = None
        ohm_idx = None
        for idx, name in enumerate(header):
            label = (name or "").strip().lower()
            if label == "date_and_time":
                time_idx = idx
            if "top centre" in label and "ohm" in label:
                ohm_idx = idx

        if time_idx is None or ohm_idx is None:
            return points

        for row in reader:
            if not row or len(row) <= max(time_idx, ohm_idx):
                continue
            t = _parse_time(row[time_idx])
            resistance = _safe_float(row[ohm_idx])
            if t is None or resistance is None:
                continue
            temp = _pt1000_lookup_temp_from_ohms(resistance, lookup)
            if temp is None:
                continue
            points.append(SeriesPoint(t=t, value=temp))

    return points


def _nearest_value(points: list[SeriesPoint], target: datetime) -> float | None:
    if not points:
        return None
    best = min(points, key=lambda p: abs((p.t - target).total_seconds()))
    if abs((best.t - target).total_seconds()) > MAX_NEAREST_DELTA_SECONDS:
        return None
    return best.value


def _stable_dwell_seconds(points: list[SeriesPoint], switch_time: datetime, setpoint_c: float, tolerance_c: float) -> float:
    pre = [point for point in points if point.t <= switch_time]
    if not pre:
        return 0.0
    pre.sort(key=lambda p: p.t)

    in_band_indices = [
        idx for idx, point in enumerate(pre) if abs(point.value - setpoint_c) <= tolerance_c
    ]
    if not in_band_indices:
        return 0.0

    end_index = in_band_indices[-1]
    if (switch_time - pre[end_index].t).total_seconds() > REQUIRE_IN_BAND_NEAR_SWITCH_SECONDS:
        return 0.0

    start_index = end_index
    while start_index - 1 >= 0:
        current = pre[start_index]
        previous = pre[start_index - 1]

        if abs(previous.value - setpoint_c) > tolerance_c:
            break

        delta_seconds = (current.t - previous.t).total_seconds()
        if delta_seconds <= 0:
            break
        slope_c_per_min = abs(current.value - previous.value) / (delta_seconds / 60.0)
        if slope_c_per_min > MAX_PLATEAU_SLOPE_C_PER_MIN:
            break

        start_index -= 1

    start_time = pre[start_index].t
    return max(0.0, (switch_time - start_time).total_seconds())


def _run_switch_times(root: Path) -> dict[str, datetime]:
    result: dict[str, datetime] = {}
    for rs422_log in _find_preferred_rs422_logs(root):
        packets = _read_all_packets(rs422_log)
        hk_packets = [pkt for pkt_type, pkt in packets if pkt_type == "HK"]
        hk_packets = [pkt for pkt in hk_packets if getattr(pkt, "TIME", None) is not None]
        if not hk_packets:
            continue

        switch_idx = _find_ob_switch_on_index(hk_packets)
        switch_idx = _first_valid_sample_after_index(hk_packets, switch_idx)
        switch_time = hk_packets[switch_idx].TIME

        run_key = _extract_sft_key(rs422_log.parent.name)
        if run_key is not None:
            result[run_key] = switch_time
    return result


def _ensure_column(ws, header: str, after_header: str) -> int:
    header_row = 1
    headers = [ws.cell(header_row, c).value for c in range(1, ws.max_column + 1)]
    for idx, value in enumerate(headers, start=1):
        if str(value).strip() == header:
            return idx

    insert_at = None
    for idx, value in enumerate(headers, start=1):
        if str(value).strip() == after_header:
            insert_at = idx + 1
            break

    if insert_at is None:
        insert_at = ws.max_column + 1

    ws.insert_cols(insert_at)
    ws.cell(header_row, insert_at).value = header
    return insert_at


def _fmt(value: float | None) -> float | None:
    if value is None:
        return None
    return round(float(value), 3)


def fill_workbook() -> None:
    eb_series = _read_eb_series(EB_CSV_PATH)
    ob_series = _read_ob_series(OB_CSV_PATH)
    rov_series = _read_rov_series(ROV_CSV_PATH)
    switch_times = _run_switch_times(THERMAL_ROOT)

    wb = load_workbook(WORKBOOK_PATH)
    ws = wb["Summary"]

    headers = {str(ws.cell(1, c).value).strip(): c for c in range(1, ws.max_column + 1)}
    temp_col = headers["Temp (°C)"]
    run_col = headers["run"]

    eb_dwell_col = _ensure_column(ws, EB_DWELL_HEADER, "Dwell Time (s)")
    ob_dwell_col = _ensure_column(ws, OB_DWELL_HEADER, EB_DWELL_HEADER)

    headers = {str(ws.cell(1, c).value).strip(): c for c in range(1, ws.max_column + 1)}
    eb_col = headers["EB Chamber Temp (°C)"]
    ob_col = headers["OB Chamber Temp(°C)"]
    rov_col = headers["ROV_THERM (°C)"]

    overwrite_runs = {"SFT1", "SFT2", "SFT3", "SFT5"}

    for row in range(2, ws.max_row + 1):
        run_text = str(ws.cell(row, run_col).value or "")
        temp_text = str(ws.cell(row, temp_col).value or "")
        run_key = _extract_sft_key(run_text)
        setpoint = _parse_setpoint(temp_text)
        if run_key is None or setpoint is None:
            continue

        switch_time = switch_times.get(run_key)
        if switch_time is None:
            continue

        eb_val = _nearest_value(eb_series, switch_time)
        ob_val = _nearest_value(ob_series, switch_time)
        rov_val = _nearest_value(rov_series, switch_time)
        eb_dwell = _stable_dwell_seconds(eb_series, switch_time, setpoint, STABILITY_TOLERANCE_C)
        ob_dwell = _stable_dwell_seconds(ob_series, switch_time, setpoint, STABILITY_TOLERANCE_C)

        if run_key in overwrite_runs:
            ws.cell(row, eb_col).value = _fmt(eb_val)
            ws.cell(row, ob_col).value = _fmt(ob_val)
            ws.cell(row, rov_col).value = _fmt(rov_val)
        else:
            if ws.cell(row, eb_col).value in (None, ""):
                ws.cell(row, eb_col).value = _fmt(eb_val)
            if ws.cell(row, ob_col).value in (None, ""):
                ws.cell(row, ob_col).value = _fmt(ob_val)
            if ws.cell(row, rov_col).value in (None, ""):
                ws.cell(row, rov_col).value = _fmt(rov_val)

        if run_key in overwrite_runs and ws.cell(row, eb_col).value is None and eb_val is not None:
            ws.cell(row, eb_col).value = _fmt(eb_val)

        ws.cell(row, eb_dwell_col).value = _fmt(eb_dwell)
        ws.cell(row, ob_dwell_col).value = _fmt(ob_dwell)

    wb.save(WORKBOOK_PATH)


if __name__ == "__main__":
    fill_workbook()
    print("Updated workbook with EB/OB/ROV chamber temperatures and dwell times")
